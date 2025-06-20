from datetime import datetime
import time
from typing import Optional, Set
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
import pandas as pd
from collections import defaultdict
from typing import Dict, Any, List

import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from service.routes.mbj_routes import get_mbj_details

EXPORT_DIR = "./exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

EXCEL_DEFECT_COLUMNS = {
    "NG Generali": "Generali",
    "NG Disall. Stringa": "Disallineamento",
    "NG Disall. Ribbon": "Disallineamento",
    "NG Saldatura": "Saldatura",
    "NG Mancanza I_Ribbon": "Mancanza Ribbon",
    "NG I_Ribbon Leadwire": "I_Ribbon Leadwire",
    "NG Macchie ECA": "Macchie ECA",
    "NG Celle Rotte": "Celle Rotte",
    "NG Altro": "Altro",
    "NG Graffio su Cella": "Graffio su Cella",
    "NG Bad Soldering": "Bad Soldering",
}

SHEET_NAMES = [
    "Metadata", "Risolutivo", "Eventi", #"MBJ", "Rework",
    "NG Generali", "NG Saldature", "NG Disall. Ribbon",
    "NG Disall. Stringa", "NG Mancanza Ribbon", "NG I_Ribbon Leadwire", "NG Macchie ECA", "NG Celle Rotte", "NG Lunghezza String Ribbon", "NG Graffio su Cella", "NG Bad Soldering", "NG Altro"
]

MBJ_FIELD_PREFIXES = {
    "Mostra Ribbon": "Ribbon-Cell",
    "Gap Orizzontali tra Celle": "GapY",
    "Gap Verticali tra Celle": "GapX",
    "Distanza Vetro-Cella": "Glass-Cell",
    "Distanza Vetro-Ribbon": "Glass-Ribbon",
    "Mostra Warnings": "Avvisi",
}


def clean_old_exports(max_age_hours: int = 2):
    now = time.time()
    for filename in os.listdir(EXPORT_DIR):
        path = os.path.join(EXPORT_DIR, filename)
        if os.path.isfile(path):
            age = now - os.path.getmtime(path)
            if age > max_age_hours * 3600:
                print(f"üóëÔ∏è Deleting old file: {filename}")
                os.remove(path)

def export_full_excel(data: dict, progress_callback=None) -> str:
    """Generate the Excel file and optionally report progress."""
    filename = f"Esportazione_PMS_{datetime.now().strftime('%d-%m-%Y.%H-%M')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    if progress_callback:
        progress_callback("init")

    wb = Workbook()

    # --- Remove default sheet if present
    default_sheet = wb.active
    if isinstance(default_sheet, Worksheet):
        wb.remove(default_sheet)

    total_modules = len(data.get("id_moduli", []))
    progress_state = {
        "callback": progress_callback,
        "current": 0,
        "total": total_modules,
        "seen": set(),
        "sheet": None,
    }

    if progress_callback:
        progress_callback("start_sheets", progress_state["current"], progress_state["total"])

    for sheet_name in SHEET_NAMES:
        func = SHEET_FUNCTIONS.get(sheet_name)
        assert func is not None, f"Sheet function not found for sheet '{sheet_name}'"

        if progress_callback:
            progress_callback(f"creating:{sheet_name}", progress_state["current"], progress_state["total"])

        progress_state["sheet"] = sheet_name
        progress_state["last_key"] = None

        ws = wb.create_sheet(title=sheet_name)
        result = func(ws, data, progress=progress_state)

        if progress_callback:
            progress_callback(f"finished:{sheet_name}", progress_state["current"], progress_state["total"])

        # Only keep sheet if result is True or function is not boolean-returning (e.g., Metadata)
        if result is False:
            wb.remove(ws)

    if progress_callback:
        progress_callback("saving")

    wb.save(filepath)

    if progress_callback:
        progress_callback("done")

    return filename

def autofit_columns(ws, align_center_for: Optional[Set[str]] = None):
    """
    Adjust column widths based on header and cell values, and apply alignment.

    :param ws: The worksheet to modify.
    :param align_center_for: Optional set of column headers that should be center-aligned.
    """
    if align_center_for is None:
        align_center_for = set()

    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        col_letter = get_column_letter(col_idx)
        header = ws[f"{col_letter}1"].value
        max_len = len(str(header)) if header else 0

        for cell in column_cells[1:]:  # Skip header
            val_len = len(str(cell.value)) if cell.value else 0
            max_len = max(max_len, val_len)

            if header in align_center_for:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions[col_letter].width = max_len + 2

def metadata_sheet(ws, data: dict, progress=None):
    from datetime import datetime

    current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    productions = data.get("productions", [])
    id_moduli = data.get("id_moduli", [])
    filters = data.get("filters", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    ws.append(["üìù METADATI ESPORTAZIONE"])
    ws.append([])
    ws.append(["Data e ora esportazione:", current_time])
    ws.append(["Numero totale moduli esportati:", len(id_moduli)])
    ws.append(["Numero totale eventi esportati:", len(productions)])

    # ‚ûï Breakdown by adjusted esito logic
    good = 0
    no_good = 0
    ok_op = 0
    not_checked = 0
    escluso = 0
    in_production = 0

    for p in productions:
        raw_esito = p.get("esito")
        cycle_time_obj = p.get("cycle_time")

        # Convert to seconds if possible
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except:
                pass

        # Classification logic (similar to map_esito)
        if raw_esito == 6:
            no_good += 1
        elif raw_esito == 1:
            if cycle_seconds is not None and cycle_seconds < min_cycle_threshold:
                not_checked += 1
            else:
                good += 1
        elif raw_esito == 5:
            if cycle_seconds is not None and cycle_seconds < min_cycle_threshold:
                not_checked += 1
            else:
                ok_op += 1
        elif raw_esito == 2:
            in_production += 1
        elif raw_esito == 4:
            escluso += 1

    # Add to sheet
    ws.append(["Good:", good])
    ws.append(["Good Operatore (ReWork):", ok_op])
    ws.append(["No Good:", no_good])
    ws.append(["Not Controllati QG2:", not_checked])
    ws.append(["Escluso:", escluso])
    ws.append(["In Produzione:", in_production])

    # Threshold info
    ws.append(["Soglia Minima Tempo Ciclo (sec):", f"{min_cycle_threshold:.1f}"])
    ws.append([])

    ws.append(["Filtri Attivi"])
    if filters:
        for f in filters:
            raw_value = f.get("value", "")
            segments = [seg.strip() for seg in raw_value.split(">") if seg.strip()]
            cleaned_value = " > ".join(segments)
            ws.append([f.get("type", "Filtro"), cleaned_value])
    else:
        ws.append(["Nessun filtro applicato"])
    ws.append([])

    # Autofit
    left_align = Alignment(horizontal="left", vertical="center")
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            cell.alignment = left_align
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 4

# ---------------------------------------------------------------------------
# Helper utilities -----------------------------------------------------------
# ---------------------------------------------------------------------------

FILL_BLUE = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_QG = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light grey


# Excel helpers --------------------------------------------------------------

def _append_dataframe(
    ws: Worksheet,
    df: pd.DataFrame,
    zebra_key: str = "Module Id",
    progress: dict | None = None,
):
    """Write *df* to *ws* with a blue/white zebra pattern keyed on *zebra_key*."""
    current_key = None
    current_fill = FILL_WHITE

    # Header first
    ws.append(df.columns.tolist())

    columns = list(df.columns)

    # ü©∫ Check if zebra_key exists, else disable zebra logic
    if zebra_key in df.columns:
        zebra_idx = df.columns.get_loc(zebra_key)
    else:
        zebra_idx = None  # disable zebra striping

    for row_tuple in df.itertuples(index=False, name=None):
        if zebra_idx is not None:
            key_value = row_tuple[zebra_idx]
            if key_value != current_key:
                current_fill = FILL_BLUE if current_fill == FILL_WHITE else FILL_WHITE
                current_key = key_value
        else:
            # If zebra_idx is None, just always keep current_fill
            pass

        row_values: List[Any] = []
        for idx, val in enumerate(row_tuple):
            col = columns[idx]
            if col in {"Checkin - PMS", "Checkout - PMS"} and val:
                try:
                    row_values.append(val.strftime("%d.%m.%y %H:%M:%S"))
                except Exception:
                    row_values.append(val)
            else:
                row_values.append(val)

        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx in range(1, len(row_values) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = current_fill

        if progress and zebra_idx is not None:
            key_value = row_tuple[zebra_idx]
            if key_value not in progress.get("seen", set()):
                progress.setdefault("seen", set()).add(key_value)
                progress["current"] += 1
                cb = progress.get("callback")
                if cb:
                    cb(f"creating:{progress.get('sheet')}", progress["current"], progress["total"])

# ---------------------------------------------------------------------------
# Sheet generators -----------------------------------------------------------
# ---------------------------------------------------------------------------

def risolutivo_sheet(ws: Worksheet, data: Dict[str, Any], progress=None):
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold: float = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {o["id"]: o for o in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}

    # Map defects by production_id
    defects_by_production = {}
    for d in object_defects:
        pid = d.get("production_id")
        if pid not in defects_by_production:
            defects_by_production[pid] = []
        defects_by_production[pid].append(d)

    rows: List[Dict[str, Any]] = []
    # Initialize rework counter per (module_id, station_id)
    rework_counters = defaultdict(int)

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        station_id = prod.get("station_id")
        if not station_id:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_display_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_display_name = line_display_name.removeprefix("Linea ")
        else:
            line_display_name = "Unknown"

        station_name = f"{station_name}{line_display_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_display_name}"

        # Compute current rework count
        pair = (id_modulo, station_id)
        current_rework = rework_counters[pair]
        rework_counters[pair] += 1

        # build NG Causale from this production
        prod_defects = defects_by_production.get(production_id, [])
        ng_labels = set()
        for d in prod_defects:
            cat = d.get("category")
            if cat == "Disallineamento":
                if d.get("stringa") is not None:
                    ng_labels.add("NG Disall. Stringa")
                elif d.get("i_ribbon") is not None:
                    ng_labels.add("NG Disall. Ribbon")
            elif cat == "Generali":
                defect_type = d.get("defect_type")
                ng_labels.add(f"NG {defect_type}" if defect_type else "NG Generali")
            elif cat == "Saldatura":
                ng_labels.add("NG Saldatura")
            elif cat == "Mancanza Ribbon":
                ng_labels.add("NG Mancanza I_Ribbon")
            elif cat == "I_Ribbon Leadwire":
                ng_labels.add("NG I_Ribbon Leadwire")
            elif cat == "Macchie ECA":
                ng_labels.add("NG Macchie ECA")
            elif cat == "Celle Rotte":
                ng_labels.add("NG Celle Rotte")
            elif cat == "Lunghezza String Ribbon":
                ng_labels.add("NG Lunghezza String Ribbon")
            elif cat == "Graffio su Cella":
                ng_labels.add("NG Graffio su Cella")
            elif cat == "Bad Soldering":
                ng_labels.add("NG Bad Soldering")
            elif cat == "Altro":
                ng_labels.add("NG Altro")

        # Only for Rework stations, get the NG from the latest QG before this
        ng_labels_qg = set()
        if station and station.get("type") == "rework":
            obj_id = prod.get("object_id")
            prod_time = prod.get("start_time")

            # find latest previous QG production
            prev_qg = sorted(
                [
                    p for p in productions
                    if p.get("object_id") == obj_id
                    and p.get("start_time") < prod_time
                    and stations_by_id.get(p.get("station_id"), {}).get("type") == "qc"
                ],
                key=lambda x: x.get("start_time"),
                reverse=True,
            )
            if prev_qg:
                prev_qg_id = prev_qg[0].get("id")
                prev_qg_defects = defects_by_production.get(prev_qg_id, [])
                for d in prev_qg_defects:
                    cat = d.get("category")
                    if cat == "Disallineamento":
                        if d.get("stringa") is not None:
                            ng_labels_qg.add("NG Disall. Stringa")
                        elif d.get("i_ribbon") is not None:
                            ng_labels_qg.add("NG Disall. Ribbon")
                    elif cat == "Generali":
                        defect_type = d.get("defect_type")
                        ng_labels_qg.add(f"NG {defect_type}" if defect_type else "NG Generali")
                    elif cat == "Saldatura":
                        ng_labels_qg.add("NG Saldatura")
                    elif cat == "Mancanza Ribbon":
                        ng_labels_qg.add("NG Mancanza I_Ribbon")
                    elif cat == "I_Ribbon Leadwire":
                        ng_labels_qg.add("NG I_Ribbon Leadwire")
                    elif cat == "Macchie ECA":
                        ng_labels_qg.add("NG Macchie ECA")
                    elif cat == "Celle Rotte":
                        ng_labels_qg.add("NG Celle Rotte")
                    elif cat == "Lunghezza String Ribbon":
                        ng_labels_qg.add("NG Lunghezza String Ribbon")
                    elif cat == "Graffio su Cella":
                        ng_labels_qg.add("NG Graffio su Cella")
                    elif cat == "Bad Soldering":
                        ng_labels_qg.add("NG Bad Soldering")
                    elif cat == "Altro":
                        ng_labels_qg.add("NG Altro")

        rows.append(
            {
                "Line": line_display_name,
                "Eq - PMS": station_name,
                "Stringatrice": last_station_name,
                "Module Id": id_modulo,
                "Checkin - PMS": start_time,
                "Checkout - PMS": end_time,
                "Rework": current_rework,
                "Esito": esito,
                "Tempo Ciclo": cycle_time_str,
                "NG Causale": ";".join(sorted(ng_labels)) if ng_labels else "",
                "NG Causale (QG)": ";".join(sorted(ng_labels_qg)) if ng_labels_qg else "",
            }
        )

    df = pd.DataFrame(
        rows,
        columns=[
            "Line",
            "Eq - PMS",
            "Stringatrice",
            "Module Id",
            "Checkin - PMS",
            "Checkout - PMS",
            "Rework",
            "Esito",
            "Tempo Ciclo",
            "NG Causale",
            "NG Causale (QG)",
        ],
    )

    _append_dataframe(ws, df, progress=progress)
    autofit_columns(ws, align_center_for={"Esito", "NG Causale (QG)"})

def eventi_sheet(ws, data: dict, progress=None):
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    production_lines_by_id = {line["id"]: line for line in production_lines}

    # Count how many times each modulo passes through each station
    modulo_station_counts = defaultdict(int)
    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue
        id_modulo = obj.get("id_modulo")
        station_id = prod.get("station_id")
        if id_modulo and station_id:
            modulo_station_counts[(id_modulo, station_id)] += 1

    # Keep track of which (modulo, station) pairs we've already written
    seen_pairs = set()
    rows = []

    for prod in productions:
        object_id = prod.get("object_id")
        obj = objects_by_id.get(object_id)
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        station_id = prod.get("station_id")
        if not id_modulo or not station_id:
            continue

        pair = (id_modulo, station_id)
        if pair in seen_pairs:
            continue  # Skip duplicates

        seen_pairs.add(pair)

        modulo_event_count = modulo_station_counts.get(pair, 0)

        station = stations_by_id.get(prod.get("station_id"))

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_display_name = production_lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_display_name = line_display_name.removeprefix("Linea ")
        else:
            line_display_name = "Unknown"

        station_name = f"{station_name}{line_display_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_display_name}"

        row = {
            "Line": line_display_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Numero Eventi": modulo_event_count
        }
        rows.append(row)

    df = pd.DataFrame(rows, columns=[
        "Line", "Eq - PMS", "Stringatrice", "Module Id", "Numero Eventi"
    ])

    _append_dataframe(ws, df, progress=progress)
    autofit_columns(ws, align_center_for={"Esito", "Numero Eventi"})

def rework_sheet(ws: Worksheet, data: Dict[str, Any], progress=None):
    from collections import defaultdict

    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_id = {o["id"]: o for o in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}

    defects_by_production = defaultdict(list)
    for d in object_defects:
        pid = d.get("production_id")
        defects_by_production[pid].append(d)

    def extract_causale(defects):
        ng_labels = set()
        for d in defects:
            cat = d.get("category")
            if cat == "Disallineamento":
                if d.get("stringa") is not None:
                    ng_labels.add("NG Disall. Stringa")
                elif d.get("i_ribbon") is not None:
                    ng_labels.add("NG Disall. Ribbon")
            elif cat == "Generali":
                defect_type = d.get("defect_type")
                ng_labels.add(f"NG {defect_type}" if defect_type else "NG Generali")
            elif cat == "Saldatura":
                ng_labels.add("NG Saldatura")
            elif cat == "Mancanza Ribbon":
                ng_labels.add("NG Mancanza I_Ribbon")
            elif cat == "I_Ribbon Leadwire":
                ng_labels.add("NG I_Ribbon Leadwire")
            elif cat == "Macchie ECA":
                ng_labels.add("NG Macchie ECA")
            elif cat == "Celle Rotte":
                ng_labels.add("NG Celle Rotte")
            elif cat == "Lunghezza String Ribbon":
                ng_labels.add("NG Lunghezza String Ribbon")
            elif cat == "Graffio su Cella":
                ng_labels.add("NG Graffio su Cella")
            elif cat == "Bad Soldering":
                ng_labels.add("NG Bad Soldering")
            elif cat == "Altro":
                ng_labels.add("NG Altro")
        
        if not ng_labels:
            return ""
        if len(ng_labels) > 1:
            return "NG MIX"
        return next(iter(ng_labels))

    # üîß First collect all modules that went through ReWork
    modules_in_rework = set()
    for prod in productions:
        station = stations_by_id.get(prod.get("station_id"))
        if not station or station.get("type") != "rework":
            continue

        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if id_modulo:
            modules_in_rework.add(id_modulo)

    # 1Ô∏è‚É£ Build QG2 causale map per module only for modules that entered ReWork
    module_to_qg_causale = {}

    for prod in productions:
        station = stations_by_id.get(prod.get("station_id"))
        if not station or station.get("type") != "qc":
            continue

        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo or id_modulo not in modules_in_rework:
            continue

        defects = defects_by_production.get(prod.get("id"), [])
        causale = extract_causale(defects)
        module_to_qg_causale[id_modulo] = causale

    # 2Ô∏è‚É£ Process only rework productions
    rows = []

    for prod in productions:
        station = stations_by_id.get(prod.get("station_id"))
        if not station or station.get("type") != "rework":
            continue

        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        qg_causale = module_to_qg_causale.get(id_modulo, "")
        if not qg_causale:
            # fallback: extract from rework defects
            defects_rework = defects_by_production.get(prod.get("id"), [])
            qg_causale = extract_causale(defects_rework)

        if not qg_causale:
            qg_causale = "BLANKS"

        rows.append({
            "Module Id": id_modulo,
            "Causale NG": qg_causale
        })

    # 3Ô∏è‚É£ Aggregate

    # Count Matrici al QG2 (from module_to_qg_causale)
    qg2_counts = defaultdict(int)
    for id_modulo in modules_in_rework:
        causale = module_to_qg_causale.get(id_modulo, "")
        label = causale if causale else "BLANKS"
        qg2_counts[label] += 1

    # Count Passaggi al RWK (from rework rows)
    rwk_counts = defaultdict(int)
    for row in rows:
        causale = row['Causale NG']
        rwk_counts[causale] += 1

    all_causali = set(qg2_counts.keys()).union(set(rwk_counts.keys()))
    total_rwk = sum(rwk_counts.values())

    export_rows = []
    for causale in sorted(all_causali):
        qg2 = qg2_counts.get(causale, 0)
        rwk = rwk_counts.get(causale, 0)
        perc = (rwk / total_rwk * 100) if total_rwk > 0 else 0
        export_rows.append({
            'Causale NG': causale,
            'Matrici al QG2': qg2,
            'Passaggi al RWK': rwk,
            '%': f"{perc:.1f}%"
        })

    export_rows.append({
        'Causale NG': 'Grand Total',
        'Matrici al QG2': sum(qg2_counts.values()),
        'Passaggi al RWK': total_rwk,
        '%': "100%"
    })

    summary_df = pd.DataFrame(export_rows, columns=['Causale NG', 'Matrici al QG2', 'Passaggi al RWK', '%'])
    _append_dataframe(ws, summary_df, progress=progress)
    autofit_columns(ws, align_center_for={'%', 'Passaggi al RWK'})

def mbj_sheet(ws: Worksheet, data: Dict[str, Any], progress=None):
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    min_cycle_threshold: float = data.get("min_cycle_threshold", 3.0)
    mbj_fields: Dict[str, str] = data.get("mbj_fields", {})

    enabled_prefixes = {
        MBJ_FIELD_PREFIXES[k]
        for k, v in mbj_fields.items()
        if v and k in MBJ_FIELD_PREFIXES
    }

    objects_by_id = {o["id"]: o for o in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}

    rows: List[Dict[str, Any]] = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_display_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_display_name = line_display_name.removeprefix("Linea ")
        else:
            line_display_name = "Unknown"

        station_name = f"{station_name}{line_display_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_display_name}"

        base_row = {
            "Line": line_display_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }

        raw = get_mbj_details(id_modulo)
        mbj_values = raw if isinstance(raw, dict) else None

        if mbj_values and enabled_prefixes:
            if "Ribbon-Cell" in enabled_prefixes:
                for row_idx, ribbon in mbj_values.get("interconnection_ribbon", {}).items():
                    for side in ["left", "right"]:
                        for pos in ["top", "bottom"]:
                            val = ribbon.get(side, {}).get(pos)
                            if val is not None:
                                side_label = side.capitalize()
                                pos_label = pos.capitalize()
                                col_name = f"Ribbon-Cell [{row_idx}] {side_label} {pos_label}"
                                base_row[col_name] = val

            if "GapY" in enabled_prefixes:
                for row_idx, gaps in mbj_values.get("horizontal_cell_mm", {}).items():
                    for col_idx, val in enumerate(gaps):
                        if val is not None:
                            base_row[f"GapY [{row_idx},{col_idx}]"] = val

            if "GapX" in enabled_prefixes:
                for col_idx, gaps in mbj_values.get("vertical_cell_mm", {}).items():
                    for row_idx, val in enumerate(gaps):
                        if val is not None:
                            base_row[f"GapX [{row_idx},{col_idx}]"] = val

            if "Glass-Cell" in enabled_prefixes:
                for side in ["top", "bottom"]:
                    values = mbj_values.get("glass_cell_mm", {}).get(side, [])
                    for col_idx, val in enumerate(values):
                        if val is not None:
                            base_row[f"Glass-Cell {side.capitalize()} [{col_idx}]"] = val

            if "Glass-Ribbon" in enabled_prefixes:
                # Placeholder if needed later
                pass

            if "Avvisi" in enabled_prefixes:
                defects = mbj_values.get("cell_defects")
                if defects:
                    base_row["Avvisi"] = "; ".join(str(d) for d in defects)

        rows.append(base_row)

    df = pd.DataFrame(rows)
    _append_dataframe(ws, df, progress=progress)
    autofit_columns(ws, align_center_for={"Esito", "Tempo Ciclo"})

def ng_generali_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG Generali' sheet using pre-fetched data.
    Both QG-originated and ReWork defects show "NG", but only QG-only cells get gray fill.
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    # Build lookups
    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type")
        for p in productions
    }

    # 1) Group all QG‚Äêfound "Generali" defects by id_modulo
    qg_gen_by_modulo = {}
    for d in object_defects:
        prod_id = d.get("production_id")
        if not prod_id:
            continue
        # Only consider defects that happened in a QC station
        if station_type_by_prod.get(prod_id) != "qc":
            continue
        if d.get("category") != "Generali":
            continue

        prod = productions_by_id.get(prod_id)
        if not prod:
            continue

        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        mod_id = obj["id_modulo"]
        qg_gen_by_modulo.setdefault(mod_id, []).append(d)

    # 2) Build the header
    defect_columns = [
        "Poe Scaduto", "No Good da Bussing", "Materiale Esterno su Celle", "Passthrough al Bussing",
        "Poe in Eccesso", "Solo Poe", "Solo Vetro", "Matrice Incompleta",
        "Molteplici Bus Bar", "Test"
    ]
    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + defect_columns

    all_rows = []
    grey_cells = []

    for prod in productions:
        object_id = prod.get("object_id")
        obj = objects_by_id.get(object_id)
        if not obj:
            continue
        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_display_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_display_name = line_display_name.removeprefix("Linea ")
        else:
            line_display_name = "Unknown"

        station_name = f"{station_name}{line_display_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_display_name}"

        # 3) Collect ReWork defects for "Generali"
        rework_defects_dict = {
            d.get("defect_type")
            for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Generali"
        }

        # 4) Get QG defects for this id_modulo
        fallback_qg_list = qg_gen_by_modulo.get(id_modulo, [])
        qg_defect_types = {d.get("defect_type") for d in fallback_qg_list}

        # 5) Build final "all_defects" set (= union of QG types and ReWork types)
        all_defect_types = qg_defect_types.union(rework_defects_dict)
        if not all_defect_types:
            continue

        # 6) Build row dictionary
        row = {
            "Line": line_display_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }

        # 7) For each defect column, assign "NG" if present, and gray‚Äêfill only if QG-only.
        for col in defect_columns:
            if col in rework_defects_dict or col in qg_defect_types:
                row[col] = "NG"
                # Only gray if it is in QG set AND not overridden by ReWork
                if col in qg_defect_types and col not in rework_defects_dict:
                    grey_cells.append((len(all_rows) + 2, col))
            else:
                row[col] = ""

        all_rows.append(row)
        rows_written += 1

    # 8) Append DataFrame and apply gray fill
    if rows_written > 0:
        import pandas as pd
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Map header name ‚Üí column index
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(defect_columns + ["Esito"]))
        return True

    return False

def ng_saldature_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG Saldature' sheet using pre-fetched data.
    One row per production that has at least one 'Saldatura' defect,
    either directly or inherited from QG (colored in gray).
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    # Build lookups
    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type")
        for p in productions
    }

    # 1) Group all QG‚Äêfound ‚ÄúSaldatura‚Äù defects by id_modulo,
    #    exactly the same pattern you use in ng_disall_ribbon_sheet:
    qg_sald_by_modulo = {}
    for d in object_defects:
        prod_id = d.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        if d.get("category") != "Saldatura":
            continue

        prod = productions_by_id.get(prod_id)
        if not prod:
            continue

        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        mod_id = obj["id_modulo"]
        qg_sald_by_modulo.setdefault(mod_id, []).append(d)

    # 2) Build the exact header order
    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo",
    ] + [f"Stringa {i}" for i in range(1, 13)] + [f"Stringa {i}M" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        # Convert cycle time to seconds
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        # 3) Collect all Saldatura defects at ReWork station
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Saldatura"
        ]

        # 4) If there are none in ReWork, but this is a ReWork row,
        #    fall back to any QG‚Äêcollected Saldatura defects for that module:
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = qg_sald_by_modulo.get(id_modulo, [])

        # If we have neither, skip
        if not prod_defects and not fallback_qg_defects:
            continue

        # 5) Prepare a 24‚Äêcell buffer, one slot per header column
        saldatura_cols = [""] * 24
        grey_indexes = set()

        def _mark_ng(stringa, lato, s_ribbon, is_qg_flag=False):
            """
            stringa: the stringa number
            lato: ‚ÄúM‚Äù or something else
            s_ribbon: the s_ribbon value to show
            is_qg_flag: if True, we'll gray‚Äêhighlight later
            """
            try:
                si = int(stringa)
            except:
                return
            if not (1 <= si <= 12):
                return

            # If lato != "M": map to index 0..11
            # If lato == "M": map to index 12..23
            if lato == "M":
                idx = (si - 1) + 12
            else:
                idx = (si - 1)

            # Append ‚ÄúNG: s_ribbon;‚Äù to that cell
            text = f"NG: {s_ribbon};"
            if saldatura_cols[idx]:
                saldatura_cols[idx] += " " + text
            else:
                saldatura_cols[idx] = text

            if is_qg_flag:
                grey_indexes.add(idx)
            else:
                grey_indexes.discard(idx)

        # 6) First mark all fallback QG defects in ‚Äúgray‚Äù (is_qg_flag=True)
        for d in fallback_qg_defects:
            _mark_ng(d.get("stringa"), d.get("ribbon_lato"), d.get("s_ribbon"), is_qg_flag=True)

        # 7) Then override any of those with actual ReWork defects (is_qg_flag=False)
        for d in prod_defects:
            _mark_ng(d.get("stringa"), d.get("ribbon_lato"), d.get("s_ribbon"), is_qg_flag=False)

        # 8) Build our row dict exactly in header order
        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }
        # Fill columns ‚ÄúStringa 1‚Äù ‚Ä¶ ‚ÄúStringa 12‚Äù then ‚ÄúStringa 1M‚Äù ‚Ä¶ ‚ÄúStringa 12M‚Äù
        for i in range(1, 13):
            row[f"Stringa {i}"] = saldatura_cols[i - 1]
            row[f"Stringa {i}M"] = saldatura_cols[(i - 1) + 12]

        # 9) Record which (row, column) need gray‚Äêfill
        #    We use len(all_rows)+2 because row 1 is header, row 2 is first data.
        for idx in grey_indexes:
            col_name = header[8 + idx]
            grey_cells.append((len(all_rows) + 2, col_name))

        all_rows.append(row)
        rows_written += 1

    # 10) Dump to DataFrame + append; then gray‚Äêfill exactly as in ribbon logic
    if rows_written > 0:
        import pandas as pd
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Map header cell ‚Üí column index
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_disall_ribbon_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions}

    # Group QG Disallineamento Ribbon defects by id_modulo
    qg_disall_by_modulo = {}
    for d in object_defects:
        prod_id = d.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        if d.get("category") == "Disallineamento" and d.get("i_ribbon") is not None and d.get("ribbon_lato") in {"F", "M", "B"}:
            prod = productions_by_id.get(prod_id)
            obj = objects_by_id.get(prod.get("object_id")) if prod else None
            if obj:
                qg_disall_by_modulo.setdefault(obj["id_modulo"], []).append(d)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo",
        "Ribbon 1 F", "Ribbon 2 F", "Ribbon 3 F",
        "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M",
        "Ribbon 1 B", "Ribbon 2 B", "Ribbon 3 B"
    ]

    ribbon_map = {
        ("F", 1): 0, ("F", 2): 1, ("F", 3): 2,
        ("M", 1): 3, ("M", 2): 4, ("M", 3): 5, ("M", 4): 6,
        ("B", 1): 7, ("B", 2): 8, ("B", 3): 9,
    }

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        # Convert cycle time to seconds
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        # Local defects
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and
               d.get("category") == "Disallineamento" and
               d.get("i_ribbon") is not None and
               d.get("ribbon_lato") in ["F", "M", "B"]
        ]

        # Fallback to QG defects if ReWork has none
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = qg_disall_by_modulo.get(id_modulo, [])

        if not prod_defects and not fallback_qg_defects:
            continue

        ribbon_cols = [""] * 10
        grey_indexes = set()

        for d in fallback_qg_defects:
            lato = d.get("ribbon_lato")
            i_ribbon = d.get("i_ribbon")
            try:
                idx = ribbon_map.get((str(lato), int(i_ribbon)))  # type: ignore
                if idx is not None:
                    ribbon_cols[idx] = "NG"
                    grey_indexes.add(idx)
            except Exception:
                continue

        for d in prod_defects:
            lato = d.get("ribbon_lato")
            i_ribbon = d.get("i_ribbon")
            try:
                idx = ribbon_map.get((str(lato), int(i_ribbon)))  # type: ignore
                if idx is not None:
                    ribbon_cols[idx] = "NG"
                    if idx in grey_indexes:
                        grey_indexes.remove(idx)
            except Exception:
                continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }

        # Fill in Ribbon values
        for i, col in enumerate([
            "Ribbon 1 F", "Ribbon 2 F", "Ribbon 3 F",
            "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M",
            "Ribbon 1 B", "Ribbon 2 B", "Ribbon 3 B"
        ]):
            row[col] = ribbon_cols[i]

        for idx in grey_indexes:
            grey_cells.append((len(all_rows) + 2, list(header)[8 + idx]))  # +2 to account for header

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Apply gray background to QG-only defects
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_disall_stringa_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions}

    # Group QG Disallineamento defects by id_modulo
    qg_disall_by_modulo = {}
    for d in object_defects:
        prod_id = d.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod.get("object_id")) if prod else None
        if obj and d.get("category") == "Disallineamento" and d.get("stringa"):
            qg_disall_by_modulo.setdefault(obj["id_modulo"], []).append(d)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        # Convert cycle time to seconds
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        # Get local defects
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and
               d.get("category") == "Disallineamento" and
               d.get("stringa") is not None
        ]

        # If Rework has no defects, fallback to QG
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = qg_disall_by_modulo.get(id_modulo, [])

        if not prod_defects and not fallback_qg_defects:
            continue

        stringa_cols = [""] * 12
        grey_indexes = set()

        for defect in fallback_qg_defects:
            stringa_num = defect.get("stringa")
            try:
                idx = int(stringa_num)
                if 1 <= idx <= 12:
                    stringa_cols[idx - 1] = "NG"
                    grey_indexes.add(idx - 1)
            except Exception:
                continue

        for defect in prod_defects:
            stringa_num = defect.get("stringa")
            try:
                idx = int(stringa_num)
                if 1 <= idx <= 12:
                    stringa_cols[idx - 1] = "NG"
                    if idx - 1 in grey_indexes:
                        grey_indexes.remove(idx - 1)
            except Exception:
                continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str
        }
        for i in range(12):
            row[f"Stringa {i + 1}"] = stringa_cols[i]

        for idx in grey_indexes:
            grey_cells.append((len(all_rows) + 2, f"Stringa {idx + 1}"))  # row+2 for header offset

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Apply gray background to QG-inherited defects
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_mancanza_ribbon_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Collect QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo",
        "Ribbon 1 F", "Ribbon 2 F", "Ribbon 3 F",
        "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M",
        "Ribbon 1 B", "Ribbon 2 B", "Ribbon 3 B"
    ]

    ribbon_map = {
        ("F", 1): 0, ("F", 2): 1, ("F", 3): 2,
        ("M", 1): 3, ("M", 2): 4, ("M", 3): 5, ("M", 4): 6,
        ("B", 1): 7, ("B", 2): 8, ("B", 3): 9,
    }

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Mancanza Ribbon"
        ]

        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Mancanza Ribbon"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        ribbon_cols = [""] * 10
        grey_map = set()

        for defect in fallback_qg_defects:
            lato = defect.get("ribbon_lato")
            i_ribbon = defect.get("i_ribbon")
            try:
                idx = ribbon_map.get((str(lato), int(i_ribbon))) # type: ignore
                if idx is not None:
                    ribbon_cols[idx] = "NG"
                    grey_map.add(idx)
            except Exception:
                continue

        for defect in prod_defects:
            lato = defect.get("ribbon_lato")
            i_ribbon = defect.get("i_ribbon")
            try:
                idx = ribbon_map.get((str(lato), int(i_ribbon))) # type: ignore
                if idx is not None:
                    ribbon_cols[idx] = "NG"
                    if idx in grey_map:
                        grey_map.remove(idx)
            except Exception:
                continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
            "Ribbon 1 F": ribbon_cols[0],
            "Ribbon 2 F": ribbon_cols[1],
            "Ribbon 3 F": ribbon_cols[2],
            "Ribbon 1 M": ribbon_cols[3],
            "Ribbon 2 M": ribbon_cols[4],
            "Ribbon 3 M": ribbon_cols[5],
            "Ribbon 4 M": ribbon_cols[6],
            "Ribbon 1 B": ribbon_cols[7],
            "Ribbon 2 B": ribbon_cols[8],
            "Ribbon 3 B": ribbon_cols[9],
        }

        for idx in grey_map:
            col_name = header[8 + idx]  # first 8 cols before ribbon cols
            grey_cells.append((len(all_rows) + 2, col_name))

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Apply grey background to QG-inherited defects
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False
    
def ng_iribbon_leadwire_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG I_Ribbon Leadwire' sheet using pre-fetched data.
    Includes ReWork rows and shows QG-inherited defects in gray.
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {s["id"]: s for s in stations}
    lines_by_id = {l["id"]: l for l in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Collect QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo",
        "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M"
    ]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "I_Ribbon Leadwire"
        ]
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "I_Ribbon Leadwire"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        ribbon_cols = [""] * 4
        grey_map = set()
        ribbon_map = {1: 0, 2: 1, 3: 2, 4: 3}

        for d in fallback_qg_defects:
            lato = d.get("ribbon_lato")
            i_ribbon = d.get("i_ribbon")
            if lato == "M":
                try:
                    idx = ribbon_map.get(int(i_ribbon))
                    if idx is not None:
                        ribbon_cols[idx] = "NG"
                        grey_map.add(idx)
                except Exception:
                    continue

        for d in prod_defects:
            lato = d.get("ribbon_lato")
            i_ribbon = d.get("i_ribbon")
            if lato == "M":
                try:
                    idx = ribbon_map.get(int(i_ribbon))
                    if idx is not None:
                        ribbon_cols[idx] = "NG"
                        if idx in grey_map:
                            grey_map.remove(idx)  # override QG mark
                except Exception:
                    continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Ribbon 1 M": ribbon_cols[0],
            "Ribbon 2 M": ribbon_cols[1],
            "Ribbon 3 M": ribbon_cols[2],
            "Ribbon 4 M": ribbon_cols[3],
        }

        for idx in grey_map:
            grey_cells.append((len(all_rows) + 2, f"Ribbon {idx + 1} M"))  # +2: account for header

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Apply grey background to QG-inherited defects
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_macchie_eca_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG Macchie ECA' sheet using pre-fetched data.
    Includes ReWork productions and marks inherited QG defects in gray.
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Collect QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Macchie ECA"
        ]
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Macchie ECA"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        found_stringa = {}
        for d in fallback_qg_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "QG"
        for d in prod_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= 12:
                found_stringa[idx] = "CURRENT"

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str
        }

        for i in range(1, 13):
            if found_stringa.get(i):
                row[f"Stringa {i}"] = "NG"
                if found_stringa[i] == "QG":
                    grey_cells.append((len(all_rows) + 2, f"Stringa {i}"))  # +2 because of header
            else:
                row[f"Stringa {i}"] = ""

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Highlight QG-inherited NG cells
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_bad_soldering_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG Bad Soldering' sheet using pre-fetched data.
    Includes ReWork productions and marks inherited QG defects in gray.
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Collect QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Bad Soldering"
        ]
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Bad Soldering"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        found_stringa = {}
        for d in fallback_qg_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "QG"
        for d in prod_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "CURRENT"

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str
        }

        for i in range(1, 13):
            if found_stringa.get(i):
                row[f"Stringa {i}"] = "NG"
                if found_stringa[i] == "QG":
                    grey_cells.append((len(all_rows) + 2, f"Stringa {i}"))  # +2 accounts for header row
            else:
                row[f"Stringa {i}"] = ""

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Highlight QG-inherited NG cells
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_celle_rotte_sheet(ws, data: dict, progress=None) -> bool:
    """
    Generate the 'NG Celle Rotte' sheet using pre-fetched data.
    Includes QG defects shown in ReWork step with gray background.
    """
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Group QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Celle Rotte"
        ]
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Celle Rotte"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        found_stringa = {}
        for d in fallback_qg_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "QG"
        for d in prod_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "CURRENT"

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str
        }

        for i in range(1, 13):
            if found_stringa.get(i):
                row[f"Stringa {i}"] = "NG"
                if found_stringa[i] == "QG":
                    grey_cells.append((len(all_rows) + 2, f"Stringa {i}"))  # +2 = header + 1-based row
            else:
                row[f"Stringa {i}"] = ""

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Gray background for QG defects
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True

    return False

def ng_lunghezza_string_ribbon_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Group QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        # Convert cycle time
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Lunghezza String Ribbon"
        ]
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Lunghezza String Ribbon"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        found_stringa = {}
        for d in fallback_qg_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "QG"
        for d in prod_defects:
            idx = d.get("stringa")
            if isinstance(idx, int) and 1 <= idx <= 12:
                found_stringa[idx] = "CURRENT"

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }

        for i in range(1, 13):
            if found_stringa.get(i):
                row[f"Stringa {i}"] = "NG"
                if found_stringa[i] == "QG":
                    grey_cells.append((len(all_rows) + 2, f"Stringa {i}"))
            else:
                row[f"Stringa {i}"] = ""

        all_rows.append(row)
        rows_written += 1

    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Gray background for QG cells
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True
    return False

def ng_graffio_su_cella_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}

    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type")
        for p in productions
    }

    # id_modulo ‚Üí QG defects
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]

    rows = []
    grey_cells = []  # List of (row_index, column_name)

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue
        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")
        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass
        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Graffio su Cella"
        ]

        fallback_qg_defects = []
        if is_rework and not prod_defects:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Graffio su Cella"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        # Map of stringa ‚Üí origin
        stringa_origin = {}
        for d in fallback_qg_defects:
            try:
                i = int(d.get("stringa"))
                if 1 <= i <= 12:
                    stringa_origin[i] = "QG"
            except: continue
        for d in prod_defects:
            try:
                i = int(d.get("stringa"))
                if 1 <= i <= 12:
                    stringa_origin[i] = "CURRENT"
            except: continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }
        for i in range(1, 13):
            label = f"Stringa {i}"
            if i in stringa_origin:
                row[label] = "NG"
                if stringa_origin[i] == "QG":
                    grey_cells.append((len(rows) + 2, label))  # +2 for Excel row offset (header + 1-based)
            else:
                row[label] = ""

        rows.append(row)

    if not rows:
        return False

    # Zebra append
    _append_dataframe(ws, pd.DataFrame(rows, columns=header), zebra_key="Module Id", progress=progress)

    # Apply grey fill
    header_map = {cell.value: cell.column for cell in ws[1]}  # Column name to index
    for row_idx, col_name in grey_cells:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

    autofit_columns(ws, align_center_for=set(header))
    return True

def ng_altro_sheet(ws, data: dict, progress=None) -> bool:
    rows_written = 0
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    min_cycle_threshold = data.get("min_cycle_threshold", 3.0)

    objects_by_id = {obj["id"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}
    productions_by_id = {p["id"]: p for p in productions}
    station_type_by_prod = {
        p["id"]: stations_by_id.get(p["station_id"], {}).get("type") for p in productions
    }

    # Step 1: gather all unique "Altro" extra_data values
    unique_altro_descriptions = sorted({
        d.get("extra_data", "").strip()
        for d in object_defects
        if d.get("category") == "Altro" and d.get("extra_data")
    })

    # Step 2: QG defects by id_modulo
    qg_defects_by_modulo = {}
    for defect in object_defects:
        prod_id = defect.get("production_id")
        if not prod_id or station_type_by_prod.get(prod_id) != "qc":
            continue
        prod = productions_by_id.get(prod_id)
        obj = objects_by_id.get(prod["object_id"]) if prod else None
        if obj:
            qg_defects_by_modulo.setdefault(obj["id_modulo"], []).append(defect)

    # Step 3: Build rows and track grey cells
    header = [
        "Line", "Eq - PMS", "Stringatrice", "Module Id",
        "Checkin - PMS", "Checkout - PMS", "Esito", "Tempo Ciclo"
    ] + unique_altro_descriptions

    all_rows = []
    grey_cells = []

    for prod in productions:
        obj = objects_by_id.get(prod.get("object_id"))
        if not obj:
            continue

        id_modulo = obj.get("id_modulo")
        if not id_modulo:
            continue

        production_id = prod.get("id")
        start_time = prod.get("start_time")
        end_time = prod.get("end_time")
        cycle_time_obj = prod.get("cycle_time")
        cycle_time_str = str(cycle_time_obj or "")

        cycle_seconds = None
        if cycle_time_obj:
            try:
                h, m, s = map(float, str(cycle_time_obj).split(":"))
                cycle_seconds = h * 3600 + m * 60 + s
            except Exception:
                pass

        esito = map_esito(prod.get("esito"), cycle_seconds, min_cycle_threshold)

        station = stations_by_id.get(prod.get("station_id"))
        is_rework = (station.get("type") == "rework") if station else False

        station_name = station.get("name", "Unknown") if station else "Unknown"

        if station:
            line_name = lines_by_id.get(station.get("line_id"), {}).get("display_name", "Unknown")
            line_name = line_name.removeprefix("Linea ")
        else:
            line_name = "Unknown"

        station_name = f"{station_name}{line_name}"


        last_station_id = prod.get("last_station_id")
        last_station_name = (
            stations_by_id.get(last_station_id, {}).get("name", "N/A")
            if last_station_id
            else "N/A"
        )
        last_station_name = f"{last_station_name}{line_name}"

        # Fetch ReWork defects first
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Altro"
        ]
        # Fallback to QG if ReWork has none
        fallback_qg_defects = []
        if not prod_defects and is_rework:
            fallback_qg_defects = [
                d for d in qg_defects_by_modulo.get(id_modulo, [])
                if d.get("category") == "Altro"
            ]

        if not prod_defects and not fallback_qg_defects:
            continue

        row = {
            "Line": line_name,
            "Eq - PMS": station_name,
            "Stringatrice": last_station_name,
            "Module Id": id_modulo,
            "Checkin - PMS": start_time,
            "Checkout - PMS": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time_str,
        }

        # Track which fields are QG
        found_descriptions = {}
        for d in fallback_qg_defects:
            desc = d.get("extra_data", "").strip()
            if desc:
                found_descriptions[desc] = "QG"
        for d in prod_defects:
            desc = d.get("extra_data", "").strip()
            if desc:
                found_descriptions[desc] = "CURRENT"

        for desc in unique_altro_descriptions:
            if found_descriptions.get(desc):
                row[desc] = "NG"
                if found_descriptions[desc] == "QG":
                    grey_cells.append((len(all_rows) + 2, desc))  # +2 offset for Excel
            else:
                row[desc] = ""

        all_rows.append(row)
        rows_written += 1

    # Step 4: Export + Zebra + Grey
    if rows_written > 0:
        df = pd.DataFrame(all_rows, columns=header)
        _append_dataframe(ws, df, zebra_key="Module Id", progress=progress)

        # Grey fill inherited QG cells
        header_map = {cell.value: cell.column for cell in ws[1]}
        for row_idx, col_name in grey_cells:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx).fill = FILL_QG

        autofit_columns(ws, align_center_for=set(header))
        return True
    return False

# --- Mapping sheet names to functions ---
SHEET_FUNCTIONS = {
    "Metadata": metadata_sheet,
    "Risolutivo": risolutivo_sheet,
    "Eventi": eventi_sheet,
    #"Rework": rework_sheet,
    #"MBJ": mbj_sheet,
    "NG Generali": ng_generali_sheet,
    "NG Saldature": ng_saldature_sheet,
    "NG Disall. Ribbon": ng_disall_ribbon_sheet,
    "NG Disall. Stringa": ng_disall_stringa_sheet,
    "NG Mancanza Ribbon": ng_mancanza_ribbon_sheet,
    "NG I_Ribbon Leadwire": ng_iribbon_leadwire_sheet,
    "NG Macchie ECA": ng_macchie_eca_sheet,
    "NG Celle Rotte": ng_celle_rotte_sheet,
    "NG Lunghezza String Ribbon": ng_lunghezza_string_ribbon_sheet,
    "NG Graffio su Cella": ng_graffio_su_cella_sheet,
    "NG Bad Soldering": ng_bad_soldering_sheet,
    "NG Altro": ng_altro_sheet,
}

def map_esito(value: Optional[int], cycle_seconds: Optional[float] = None, threshold: float = 3.0) -> str:
    if value == 1:  # G
        if cycle_seconds is not None and cycle_seconds < threshold:
            return "NC"
        return "G"
    elif value == 2:
        return "In Produzione"
    elif value == 4:
        return "Escluso"
    elif value == 5:
        return "G Operatore"
    elif value == 6:
        return "NG"
    return "N/A"
