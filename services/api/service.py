import asyncio
import configparser
import datetime
from datetime import datetime, timedelta
import json
import logging
import os
import time
from fastapi import Body, FastAPI, Form, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect, Request, BackgroundTasks, File
import base64
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from controllers.plc import PLCConnection
import uvicorn
from typing import AsyncGenerator, Optional, Set
from contextlib import asynccontextmanager
from pymysql.cursors import DictCursor
import pymysql
import re
from fastapi.staticfiles import StaticFiles
from typing import Optional, Union, Dict, List
from fastapi.responses import FileResponse
import pandas as pd
import os
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


# ---------------- CONFIG & GLOBALS ----------------
CHANNELS = {
    "Linea1": {
        "M308": {
            "trigger": {"db": 19606, "byte": 0, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 2, "length": 20},
            "id_utente": {"db": 19606, "byte": 24, "length": 20},
            "fine_buona": {"db": 19606, "byte": 0, "bit": 6},
            "fine_scarto": {"db": 19606, "byte": 0, "bit": 7},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 0},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 3},
            "stringatrice": {"db": 19606, "byte": 46, "length": 5},
        },
        "M309": {
            "trigger": {"db": 19606, "byte": 48, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 50, "length": 20},
            "id_utente": {"db": 19606, "byte": 72, "length": 20},
            "fine_buona": {"db": 19606, "byte": 48, "bit": 6},
            "fine_scarto": {"db": 19606, "byte": 48, "bit": 7},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 1},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 4},
            "stringatrice": {"db": 19606, "byte": 94, "length": 5},
        },
    "M326": {
            "trigger": {"db": 19606, "byte": 96, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 98, "length": 20},
            "id_utente": {"db": 19606, "byte": 120, "length": 20},
            "fine_buona": {"db": 19606, "byte": 96, "bit": 4},
            "fine_scarto": {"db": 19606, "byte": 96, "bit": 5},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 2},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 5},
            "stringatrice": {"db": 19606, "byte": 142, "length": 5},
        },
    },
    "Linea2": {
        "M308": {
            "trigger": {"db": 19606, "byte": 0, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 2, "length": 20},
            "id_utente": {"db": 19606, "byte": 24, "length": 20},
            "fine_buona": {"db": 19606, "byte": 0, "bit": 6},
            "fine_scarto": {"db": 19606, "byte": 0, "bit": 7},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 0},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 3},
            "stringatrice": {"db": 19606, "byte": 46, "length": 5},
        },
        "M309": {
            "trigger": {"db": 19606, "byte": 48, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 50, "length": 20},
            "id_utente": {"db": 19606, "byte": 72, "length": 20},
            "fine_buona": {"db": 19606, "byte": 48, "bit": 6},
            "fine_scarto": {"db": 19606, "byte": 48, "bit": 7},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 1},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 4},
            "stringatrice": {"db": 19606, "byte": 94, "length": 5},
        },
    "M326": {
            "trigger": {"db": 19606, "byte": 96, "bit": 4},
            "id_modulo": {"db": 19606, "byte": 98, "length": 20},
            "id_utente": {"db": 19606, "byte": 120, "length": 20},
            "fine_buona": {"db": 19606, "byte": 96, "bit": 4},
            "fine_scarto": {"db": 19606, "byte": 96, "bit": 5},
            "esito_scarto_compilato": {"db": 19606, "byte": 144, "bit": 2},
            "pezzo_salvato_su_DB_con_inizio_ciclo": {"db": 19606, "byte": 144, "bit": 5},
            "stringatrice": {"db": 19606, "byte": 142, "length": 5},
        },
    },
}

ISSUE_TREE = {
    "Dati": {
        "Esito": {
            "Esito_Scarto": {
                "Difetti": {
                    "Generali": {
                        "Non Lavorato Poe Scaduto": {},
                        "Non Lavorato da Telecamere": {},
                        "Materiale Esterno su Celle": {},
                        "Bad Soldering": {}
                    },
                    "Saldatura": {
                        f"Stringa[{i}]": {
                            f"Pin[{j}]": {
                                "F": None,
                                "M": None,
                                "B": None
                            } for j in range(1, 21)
                        } for i in range(1, 13)
                    },
                   "Disallineamento": {
                        **{f"Stringa[{i}]": None for i in range(1, 13)},
                        **{f"Ribbon[{i}]": {
                            "F": None,
                            "M": None,
                            "B": None
                        } for i in range(1, 11)}
                    },
                    "Mancanza Ribbon": {
                        f"Ribbon[{i}]": {
                            "F": None,
                            "M": None,
                            "B": None
                        } for i in range(1, 11)
                    },
                    "Macchie ECA": {
                        f"Stringa[{i}]": None for i in range(1, 13)
                    },
                    "Celle Rotte": {
                        f"Stringa[{i}]": None for i in range(1, 13)
                    },
                     "Lunghezza String Ribbon": {
                        f"Stringa[{i}]": None for i in range(1, 13)
                    },
                }
            }
        }
    }
}

TEMP_STORAGE_PATH = os.path.join("C:/IX-Monitor", "temp_data.json")

# These globals are used to track WebSocket subscriptions and PLC subscription state.
subscriptions = {}
# Temporary store for trigger timestamps
trigger_timestamps = {}
incomplete_productions = {}  # Tracks production_id per station (e.g., "Linea1.M308")
stop_threads = {}
passato_flags = {}

# Then during startup:
for line in CHANNELS:
    for station in CHANNELS[line]:
        key = f"{line}.{station}"
        stop_threads[key] = False
        passato_flags[key] = False


plc_connections = {} 

# GLOBAL
mysql_connection = None
# In-memory session history for chat
user_sessions = {} 

SESSION_TIMEOUT = 600  # 600 seconds = 10 minutes

def get_channel_config(line_name: str, channel_id: str):
    return CHANNELS.get(line_name, {}).get(channel_id)


# ---------------- LIFESPAN ----------------
@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    global mysql_connection
    mysql_connection = pymysql.connect(
        host="localhost",
        user="root",
        password="Master36!",
        database="ix_monitor",
        port=3306,
        cursorclass=DictCursor,
        autocommit=False
    )
    print("🟢 MySQL connected!")

    line_configs = load_station_configs("C:/IX-Monitor/stations.ini")

    for line, config in line_configs.items():
        plc_ip = config["PLC"]["IP"]
        plc_slot = config["PLC"]["SLOT"]

        for station in config["stations"]:
            plc_conn = PLCConnection(ip_address=plc_ip, slot=plc_slot, status_callback=make_status_callback(station))
            plc_connections[f"{line}.{station}"] = plc_conn
            asyncio.create_task(background_task(plc_conn, f"{line}.{station}"))
            print(f"🚀 Background task created for {line}.{station}")

    yield

    mysql_connection.close()
    print("🔴 MySQL disconnected.")

app = FastAPI(lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust for production use
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/images", StaticFiles(directory="C:/IX-Monitor/images"), name="images")

async def send_initial_state(websocket: WebSocket, channel_id: str, plc_connection: PLCConnection, line_name: str):
    paths = get_channel_config(line_name, channel_id)
    
    if paths is None:
        print(f"❌ Invalid config for line={line_name}, channel={channel_id}")
        await websocket.send_json({"error": "Invalid line/channel combination"})
        return

    # ✅ Use `paths` for all access now (already scoped)
    trigger_conf = paths["trigger"]
    trigger_value = await asyncio.to_thread(
        plc_connection.read_bool,
        trigger_conf["db"], trigger_conf["byte"], trigger_conf["bit"]
    )

    object_id = ""
    stringatrice = ""
    outcome = None
    issues_submitted = False

    if trigger_value:
        id_mod_conf = paths["id_modulo"]
        object_id = await asyncio.to_thread(
            plc_connection.read_string,
            id_mod_conf["db"], id_mod_conf["byte"], id_mod_conf["length"]
        )

        # 🔁 FIX THIS LINE – it was still accessing CHANNELS by only channel_id!
        str_conf = paths["stringatrice"]  # ✅ Use `paths` here
        values = [
            await asyncio.to_thread(plc_connection.read_bool, str_conf["db"], str_conf["byte"], i)
            for i in range(str_conf["length"])
        ]

        if not any(values):
            values[0] = True

        stringatrice_index = values.index(True) + 1
        stringatrice = str(stringatrice_index)

        fine_buona_conf = paths["fine_buona"]
        fine_buona = await asyncio.to_thread(
            plc_connection.read_bool,
            fine_buona_conf["db"], fine_buona_conf["byte"], fine_buona_conf["bit"]
        )

        fine_scarto_conf = paths["fine_scarto"]
        fine_scarto = await asyncio.to_thread(
            plc_connection.read_bool,
            fine_scarto_conf["db"], fine_scarto_conf["byte"], fine_scarto_conf["bit"]
        )

        if fine_buona:
            outcome = "buona"
        elif fine_scarto:
            outcome = "scarto"

        esito_conf = paths["esito_scarto_compilato"]
        issues_value = await asyncio.to_thread(
            plc_connection.read_bool,
            esito_conf["db"], esito_conf["byte"], esito_conf["bit"]
        )
        issues_submitted = issues_value is True

    await websocket.send_json({
        "trigger": trigger_value,
        "objectId": object_id,
        "stringatrice": stringatrice,
        "outcome": outcome,
        "issuesSubmitted": issues_submitted
    })

async def broadcast(line_name: str, channel_id: str, message: dict):
    key = f"{line_name}.{channel_id}"
    for ws in list(subscriptions.get(key, [])):
        try:
            await ws.send_json(message)
        except Exception:
            subscriptions[key].remove(ws)

    # Also broadcast to per-line summary, if any
    summary_key = f"{line_name}.summary"
    for ws in list(subscriptions.get(summary_key, [])):
        try:
            await ws.send_json(message)
        except Exception:
            subscriptions[summary_key].remove(ws)

async def scan_issues(node, base_path):
    selected = []
    children = await node.get_children()

    for child in children:
        browse_name = await child.read_browse_name()
        name = browse_name.Name
        full_path = f"{base_path}.{name}"

        try:
            value = await child.read_value()
        except:
            value = None

        # If not a boolean, assume it's a folder/struct and scan deeper
        if not isinstance(value, bool):
            deeper = await scan_issues(child, full_path)
            selected.extend(deeper)
        else:
            print(f"➡️ {full_path} = {value}")
            if value is True:
                selected.append(full_path)
                print(f"✅ Selected: {full_path}")

    return selected

def load_temp_data():
    # Ensure directory exists
    os.makedirs(os.path.dirname(TEMP_STORAGE_PATH), exist_ok=True)
    
    if not os.path.exists(TEMP_STORAGE_PATH):
        # Create an empty JSON file if it doesn't exist
        with open(TEMP_STORAGE_PATH, "w") as file:
            json.dump([], file, indent=4)
        return []

    # If it exists, load normally
    with open(TEMP_STORAGE_PATH, "r") as file:
        try:
            return json.load(file)
        except json.JSONDecodeError:
            # If the file is corrupted or empty
            return []

def save_temp_data(data):
    # Ensure directory exists before saving
    os.makedirs(os.path.dirname(TEMP_STORAGE_PATH), exist_ok=True)
    with open(TEMP_STORAGE_PATH, "w") as file:
        json.dump(data, file, indent=4)
        
def get_latest_issues(line_name: str, channel_id: str):
    """
    Returns the issues list for the given station.
    Searches the temporary storage for the latest entry with matching line_name and channel_id.
    """
    temp_data = load_temp_data()
    for entry in reversed(temp_data):
        if (
            entry.get("line_name") == line_name and
            entry.get("channel_id") == channel_id
        ):
            return entry.get("issues", [])
    return []

def remove_temp_issues(line_name, channel_id, object_id):
    temp_data = load_temp_data()
    filtered_data = [
        entry for entry in temp_data
        if not (
            entry.get("line_name") == line_name and
            entry.get("channel_id") == channel_id and
            entry.get("object_id") == object_id
        )
    ]
    save_temp_data(filtered_data)
    print(f"🗑️ Removed temp issue for {line_name}.{channel_id} - {object_id}")

def issue_matches_any(issues, pattern):
    return any(re.search(pattern, issue) for issue in issues)

async def on_trigger_change(plc_connection: PLCConnection, line_name: str, channel_id: str, node, val, data):
    if not isinstance(val, bool):
        return

    full_id = f"{line_name}.{channel_id}"
    paths = get_channel_config(line_name, channel_id)
    if not paths:
        print(f"❌ Config not found for {full_id}")
        return

    if val:
        print(f"🟡 Inizio Ciclo on {full_id} TRUE ...")
        trigger_timestamps.pop(full_id, None)

        # Write FALSE to esito_scarto_compilato.
        esito_conf = paths["esito_scarto_compilato"]
        await asyncio.to_thread(plc_connection.write_bool, esito_conf["db"], esito_conf["byte"], esito_conf["bit"], False)

        # Write FALSE to pezzo_salvato_su_DB_con_inizio_ciclo.
        pezzo_conf = paths["pezzo_salvato_su_DB_con_inizio_ciclo"]
        await asyncio.to_thread(plc_connection.write_bool, pezzo_conf["db"], pezzo_conf["byte"], pezzo_conf["bit"], False)

        # Read initial values.
        id_mod_conf = paths["id_modulo"]
        object_id = await asyncio.to_thread(plc_connection.read_string, id_mod_conf["db"], id_mod_conf["byte"], id_mod_conf["length"])

        str_conf = paths["stringatrice"]
        values = [await asyncio.to_thread(plc_connection.read_bool, str_conf["db"], str_conf["byte"], i) for i in range(str_conf["length"])]
        if not any(values):
            values[0] = True
        stringatrice_index = values.index(True) + 1
        stringatrice = str(stringatrice_index)

        issues_value = await asyncio.to_thread(plc_connection.read_bool, esito_conf["db"], esito_conf["byte"], esito_conf["bit"])
        issues_submitted = issues_value is True

        trigger_timestamps[full_id] = datetime.now()
        print(f"trigger_timestamps[{full_id}]: {trigger_timestamps[full_id]}")

        # Read the initial data from the PLC using read_data.
        data_inizio = trigger_timestamps[full_id]
        initial_data = await read_data(plc_connection, line_name, channel_id, richiesta_ok=False, richiesta_ko=False, data_inizio=data_inizio)
        # Insert the initial production record (with esito = 2) and store the production_id.
        prod_id = await insert_initial_production_data(initial_data, channel_id, mysql_connection)
        if prod_id:
            incomplete_productions[full_id] = prod_id

        # Write TRUE to pezzo_salvato_su_DB_con_inizio_ciclo.
        await asyncio.to_thread(plc_connection.write_bool, pezzo_conf["db"], pezzo_conf["byte"], pezzo_conf["bit"], True)

        await broadcast(line_name, channel_id, {
            "trigger": True,
            "objectId": object_id,
            "stringatrice": stringatrice,
            "outcome": None,
            "issuesSubmitted": issues_submitted
        })

    else:
        print(f"🟡 Inizio Ciclo on {full_id} FALSE ...")
        passato_flags[full_id] = False
        await broadcast(line_name, channel_id, {
            "trigger": False,
            "objectId": None,
            "stringatrice": None,
            "outcome": None,
            "issuesSubmitted": False
        })

async def read_data(
    plc_connection: PLCConnection,
    line_name: str,
    channel_id: str,
    richiesta_ko: bool,
    richiesta_ok: bool,
    data_inizio: datetime | None
):
    try:
        full_id = f"{line_name}.{channel_id}"

        if data_inizio is None:
            data_inizio = datetime.now()

        config = get_channel_config(line_name, channel_id)
        if config is None:
            return None

        data = {}

        # Read Id_Modulo
        id_mod_conf = config["id_modulo"]
        data["Id_Modulo"] = await asyncio.to_thread(
            plc_connection.read_string,
            id_mod_conf["db"], id_mod_conf["byte"], id_mod_conf["length"]
        )

        # Read Id_Utente
        id_utente_conf = config["id_utente"]
        data["Id_Utente"] = await asyncio.to_thread(
            plc_connection.read_string,
            id_utente_conf["db"], id_utente_conf["byte"], id_utente_conf["length"]
        )

        data["DataInizio"] = data_inizio
        data["DataFine"] = datetime.now()
        tempo_ciclo = data["DataFine"] - data_inizio
        data["Tempo_Ciclo"] = str(tempo_ciclo)

        # Set linea_in_lavorazione if needed
        data["Linea_in_Lavorazione"] = [line_name == "Linea1", line_name == "Linea2", False, False, False]

        # Read stringatrice bits if relevant
        str_conf = config["stringatrice"]
        values = [
            await asyncio.to_thread(plc_connection.read_bool, str_conf["db"], str_conf["byte"], i)
            for i in range(str_conf["length"])
        ]
        if not any(values):
            values[0] = True
        data["Lavorazione_Eseguita_Su_Stringatrice"] = values

        data["Compilato_Su_Ipad_Scarto_Presente"] = richiesta_ko

        return data

    except Exception as e:
        logging.error(f"[{full_id}] ❌ Error reading PLC data: {e}")
        return None

# ---------------- EXPORT EXCEL ----------------
EXPORT_DIR = "./exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

EXCEL_DEFECT_COLUMNS = {
    "NG Generali": "Generali",
    "NG Disall. Stringa": "Disallineamento",  # specific to stringa
    "NG Disall. Ribbon": "Disallineamento",   # specific to ribbon
    "NG Saldatura": "Saldatura",
    "NG Mancanza I_Ribbon": "Mancanza Ribbon",
    "NG Macchie ECA": "Macchie ECA",
    "NG Celle Rotte": "Celle Rotte",
    "NG Altro": "Altro"
}

# --- Sheet names ---
SHEET_NAMES = [
    "Metadata", "Risolutivo", "NG Generali", "NG Saldature", "NG Disall. Ribbon",
    "NG Disall. Stringa", "NG Mancanza Ribbon", "NG Macchie ECA", "NG Celle Rotte", "NG Lunghezza String Ribbon", "NG Altro"
]

def clean_old_exports(max_age_hours: int = 2):
    now = time.time()
    for filename in os.listdir(EXPORT_DIR):
        path = os.path.join(EXPORT_DIR, filename)
        if os.path.isfile(path):
            age = now - os.path.getmtime(path)
            if age > max_age_hours * 3600:
                print(f"🗑️ Deleting old file: {filename}")
                os.remove(path)

def export_full_excel(data: dict) -> str:
    filename = f"ixmonitor_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()

    # --- Remove default sheet if present
    default_sheet = wb.active
    if isinstance(default_sheet, Worksheet):
        wb.remove(default_sheet)

    for sheet_name in SHEET_NAMES:
        func = SHEET_FUNCTIONS.get(sheet_name)
        assert func is not None, f"Sheet function not found for sheet '{sheet_name}'"

        ws = wb.create_sheet(title=sheet_name)
        result = func(ws, data)

        # Only keep sheet if result is True or function is not boolean-returning (e.g., Metadata)
        if result is False:
            wb.remove(ws)

    wb.save(filepath)
    return filename


def autofit_columns(ws, align_center_for: Optional[Set[str]] = None):
    """
    Adjust column widths based on header and cell values, and apply alignment.

    :param ws: The worksheet to modify.
    :param align_center_for: Optional set of column headers that should be center-aligned.
    """
    if align_center_for is None:
        align_center_for = set()

    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        col_letter = get_column_letter(col_idx)
        header = ws[f"{col_letter}1"].value
        max_len = len(str(header)) if header else 0

        for cell in column_cells[1:]:  # Skip header
            val_len = len(str(cell.value)) if cell.value else 0
            max_len = max(max_len, val_len)

            if header in align_center_for:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions[col_letter].width = max_len + 2

def metadata_sheet(ws, data: dict):
    current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    id_moduli = data.get("id_moduli", [])
    filters = data.get("filters", [])

    ws.append(["📝 METADATI ESPORTAZIONE"])
    ws.append([])

    # General Info
    ws.append(["Data e ora esportazione:", current_time])
    ws.append(["Numero totale moduli esportati:", len(id_moduli)])
    ws.append([])

    # Filters
    ws.append(["Filtri Attivi"])

    if filters:
        for f in filters:
            raw_value = f.get("value", "")
            # Clean the value by removing empty ' > ' parts
            segments = [seg.strip() for seg in raw_value.split(">") if seg.strip()]
            cleaned_value = " > ".join(segments)
            ws.append([f.get("type", "Filtro"), cleaned_value])
    else:
        ws.append(["Nessun filtro applicato"])

    ws.append([])

    # Optional: Apply center-left alignment and auto-fit column widths
    left_align = Alignment(horizontal="left", vertical="center")

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            cell.alignment = left_align
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 4  # add some padding

def risolutivo_sheet(ws, data: dict):
    """
    Generate the 'Risolutivo' sheet using pre-fetched data.
    Expects data to include keys: "id_moduli", "objects", "productions",
    "stations", "production_lines", and "object_defects".
    """
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])
    
    # Build lookup dictionaries
    objects_by_modulo = {}
    for obj in objects:
        m = obj.get("id_modulo")
        if m is not None:
            objects_by_modulo[m] = obj

    stations_by_id = {station["id"]: station for station in stations}
    production_lines_by_id = {line["id"]: line for line in production_lines}
    
    rows = []
    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue
        object_id = obj["id"]
        
        # Retrieve productions for this object
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue
        
        # Get the most recent production (assumes start_time is a datetime)
        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")
        
        # Look up station and production line info
        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"
        line_display_name = "Unknown"
        if station and station.get("line_id"):
            pline = production_lines_by_id.get(station["line_id"])
            if pline:
                line_display_name = pline.get("display_name", "Unknown")
        
        # Look up stringatrice (from last_station_id)
        last_station_id = latest_prod.get("last_station_id")
        last_station_name = "N/A"
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")
    
        # Calculate NG flags over all defect categories related to this production
        # (Logic similar to your original get_ng_flags_for_production)
        flags = {
            "NG Generali": "",
            "NG Disall. Stringa": "",
            "NG Disall. Ribbon": "",
            "NG Saldatura": "",
            "NG Mancanza I_Ribbon": "",
            "NG Macchie ECA": "",
            "NG Celle Rotte": "",
            "NG Lunghezza String Ribbon": "",
            "NG Altro": ""
        }
        prod_defects = [d for d in object_defects if d.get("production_id") == production_id]
        for d in prod_defects:
            cat = d.get("category")
            if cat == "Disallineamento":
                if d.get("stringa") is not None:
                    flags["NG Disall. Stringa"] = "NG"
                elif d.get("i_ribbon") is not None:
                    flags["NG Disall. Ribbon"] = "NG"
            elif cat == "Generali":
                flags["NG Generali"] = "NG"
            elif cat == "Saldatura":
                flags["NG Saldatura"] = "NG"
            elif cat == "Mancanza Ribbon":
                flags["NG Mancanza I_Ribbon"] = "NG"
            elif cat == "Macchie ECA":
                flags["NG Macchie ECA"] = "NG"
            elif cat == "Celle Rotte":
                flags["NG Celle Rotte"] = "NG"
            elif cat == "Lunghezza String Ribbon":
                flags["NG Lunghezza String Ribbon"] = "NG"
            elif cat == "Altro":
                flags["NG Altro"] = "NG"
    
        row = {
            "Linea": line_display_name,
            "Stazione": station_name,
            "Stringatrice": last_station_name,
            "ID Modulo": id_modulo,
            "Data Ingresso": start_time,
            "Data Uscita": end_time,
            "Esito": esito,
            "Tempo Ciclo": cycle_time,
            **flags
        }
        rows.append(row)
    
    # Convert rows to a DataFrame
    df = pd.DataFrame(rows, columns=[
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo",
        "NG Generali", "NG Disall. Stringa", "NG Disall. Ribbon",
        "NG Saldatura", "NG Mancanza I_Ribbon", "NG Macchie ECA",
        "NG Celle Rotte", "NG Lunghezza String Ribbon", "NG Altro"
    ])
    
    # Write header and rows to the worksheet
    ws.append(df.columns.tolist())
    for _, row in df.iterrows():
        # If dates are datetime objects, format them as strings
        row_values = []
        for col in df.columns:
            val = row[col]
            if col in ["Data Ingresso", "Data Uscita"] and val:
                try:
                    row_values.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                except Exception:
                    row_values.append(val)
            else:
                row_values.append(val)
        ws.append(row_values)
    
    autofit_columns(ws, align_center_for={
    "Esito", "NG Generali", "NG Disall. Stringa", "NG Disall. Ribbon",
    "NG Saldatura", "NG Mancanza I_Ribbon", "NG Macchie ECA",
    "NG Celle Rotte", "NG Lunghezza String Ribbon", "NG Altro"
})

def ng_generali_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Generali' sheet using pre-fetched data.
    Returns True if at least one row is written (besides the header).
    """
    rows_written = 0  # ✅ Count data rows
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    production_lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo",
        "Poe Scaduto", "Non Lav. Da Telecamere",
        "Materiale Esterno su Celle", "Bad Soldering"
    ]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if obj is None:
            continue
        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue
        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_display_name = "Unknown"
        if station and station.get("line_id"):
            pline = production_lines_by_id.get(station["line_id"])
            if pline:
                line_display_name = pline.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects 
            if d.get("production_id") == production_id and d.get("category") == "Generali"
        ]
        if not prod_defects:
            continue
        rows_written += 1

        general_defects = {d.get("defect_type") for d in prod_defects}
        flag_poe = "NG" if "Non Lavorato Poe Scaduto" in general_defects else ""
        flag_tel = "NG" if "Non Lavorato da Telecamere" in general_defects else ""
        flag_materiale = "NG" if "Materiale Esterno su Celle" in general_defects else ""
        flag_bad = "NG" if "Bad Soldering" in general_defects else ""

        row = [
            line_display_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time,
            flag_poe,
            flag_tel,
            flag_materiale,
            flag_bad
        ]
        ws.append(row)

    if rows_written > 0:
        autofit_columns(ws, align_center_for={
            "Esito", "Poe Scaduto", "Non Lav. Da Telecamere",
            "Materiale Esterno su Celle", "Bad Soldering"
        })
        return True
    else:
        return False

def ng_saldature_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Saldature' sheet using pre-fetched data.
    Returns True if any rows were written, False otherwise.
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ]
    for i in range(1, 13):
        header.append(f"Stringa {i}")
        header.append(f"Stringa {i}M")
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Saldatura"
        ]
        if not prod_defects:
            continue  # No data for this row

        # Map of (stringa, lato) → list of s_ribbon
        pin_map = {}
        for defect in prod_defects:
            stringa = defect.get("stringa")
            lato = defect.get("ribbon_lato")
            s_ribbon = defect.get("s_ribbon")

            if stringa is None or lato is None or s_ribbon is None:
                continue

            try:
                key = (int(stringa), str(lato))
                pin_map.setdefault(key, []).append(str(s_ribbon))
            except (ValueError, TypeError):
                continue

        # Build saldatura cells: 24 columns (Stringa 1, 1M, ..., 12, 12M)
        saldatura_cols = [""] * 24
        for (stringa_num, lato), pins in pin_map.items():
            if not (1 <= stringa_num <= 12):
                continue
            formatted = f"NG: {';'.join(pins)};"
            col_index = (stringa_num - 1) * 2
            if lato == "M":
                col_index += 1
            saldatura_cols[col_index] = formatted

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + saldatura_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_disall_ribbon_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Disallineamento Ribbon' sheet using pre-fetched data.
    Returns True if any rows were written, False otherwise.
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo",
        "Ribbon 1 F", "Ribbon 2 F", "Ribbon 3 F",
        "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M",
        "Ribbon 1 B", "Ribbon 2 B", "Ribbon 3 B"
    ]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        # Filter defects with category = 'Disallineamento' and valid ribbon data
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and
            d.get("category") == "Disallineamento" and
            d.get("i_ribbon") is not None and
            d.get("ribbon_lato") in ["F", "M", "B"]
        ]
        if not prod_defects:
            continue

        # Initialize ribbon columns
        ribbon_cols = [""] * 10

        ribbon_map = {
            ("F", 1): 0, ("F", 2): 1, ("F", 3): 2,
            ("M", 1): 3, ("M", 2): 4, ("M", 3): 5, ("M", 4): 6,
            ("B", 1): 7, ("B", 2): 8, ("B", 3): 9,
        }

        for defect in prod_defects:
            lato = defect.get("ribbon_lato")
            i_ribbon = defect.get("i_ribbon")

            if lato in ["F", "M", "B"]:
                try:
                    idx = ribbon_map.get((str(lato), int(i_ribbon)))  # type: ignore
                    if idx is not None:
                        ribbon_cols[idx] = "NG"
                except (ValueError, TypeError):
                    continue

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + ribbon_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_disall_stringa_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Disallineamento Stringa' sheet using pre-fetched data.
    Returns True if any rows were written, False otherwise.
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and
            d.get("category") == "Disallineamento" and
            d.get("stringa") is not None
        ]
        if not prod_defects:
            continue

        stringa_cols = [""] * 12
        for defect in prod_defects:
            stringa_num = defect.get("stringa")
            if isinstance(stringa_num, int) and 1 <= stringa_num <= 12:
                stringa_cols[stringa_num - 1] = "NG"
            elif isinstance(stringa_num, str) and stringa_num.isdigit():
                index = int(stringa_num)
                if 1 <= index <= 12:
                    stringa_cols[index - 1] = "NG"

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + stringa_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_mancanza_ribbon_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Mancanza Ribbon' sheet using pre-fetched data.
    Returns True if any rows were written, False otherwise.
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo",
        "Ribbon 1 F", "Ribbon 2 F", "Ribbon 3 F",
        "Ribbon 1 M", "Ribbon 2 M", "Ribbon 3 M", "Ribbon 4 M",
        "Ribbon 1 B", "Ribbon 2 B", "Ribbon 3 B"
    ]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        # Filter defects with category = 'Mancanza Ribbon'
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Mancanza Ribbon"
        ]
        if not prod_defects:
            continue

        ribbon_cols = [""] * 10

        ribbon_map = {
            ("F", 1): 0, ("F", 2): 1, ("F", 3): 2,
            ("M", 1): 3, ("M", 2): 4, ("M", 3): 5, ("M", 4): 6,
            ("B", 1): 7, ("B", 2): 8, ("B", 3): 9,
        }

        for defect in prod_defects:
            lato = defect.get("ribbon_lato")
            i_ribbon = defect.get("i_ribbon")

            if lato in ["F", "M", "B"]:
                try:
                    idx = ribbon_map.get((str(lato), int(i_ribbon)))  # type: ignore
                    if idx is not None:
                        ribbon_cols[idx] = "NG"
                except (ValueError, TypeError):
                    continue

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + ribbon_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_macchie_eca_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Macchie ECA' sheet using pre-fetched data.
    Returns True if any rows are written (excluding the header).
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Macchie ECA"
        ]
        if not prod_defects:
            continue

        stringa_cols = [""] * 12
        for defect in prod_defects:
            stringa_num = defect.get("stringa")
            if isinstance(stringa_num, int) and 1 <= stringa_num <= 12:
                stringa_cols[stringa_num - 1] = "NG"
            elif isinstance(stringa_num, str) and stringa_num.isdigit():
                index = int(stringa_num)
                if 1 <= index <= 12:
                    stringa_cols[index - 1] = "NG"

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + stringa_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_celle_rotte_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Celle Rotte' sheet using pre-fetched data.
    Returns True if any rows are written (excluding the header).
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Celle Rotte"
        ]
        if not prod_defects:
            continue

        stringa_cols = [""] * 12
        for defect in prod_defects:
            stringa_num = defect.get("stringa")
            if isinstance(stringa_num, int) and 1 <= stringa_num <= 12:
                stringa_cols[stringa_num - 1] = "NG"
            elif isinstance(stringa_num, str) and stringa_num.isdigit():
                index = int(stringa_num)
                if 1 <= index <= 12:
                    stringa_cols[index - 1] = "NG"

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + stringa_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_lunghezza_string_ribbon_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Lunghezza String Ribbon' sheet using pre-fetched data.
    Returns True if any rows are written (excluding the header).
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ] + [f"Stringa {i}" for i in range(1, 13)]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Lunghezza String Ribbon"
        ]
        if not prod_defects:
            continue

        stringa_cols = [""] * 12
        for defect in prod_defects:
            stringa_num = defect.get("stringa")
            if isinstance(stringa_num, int) and 1 <= stringa_num <= 12:
                stringa_cols[stringa_num - 1] = "NG"
            elif isinstance(stringa_num, str) and stringa_num.isdigit():
                index = int(stringa_num)
                if 1 <= index <= 12:
                    stringa_cols[index - 1] = "NG"

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + stringa_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

def ng_altro_sheet(ws, data: dict) -> bool:
    """
    Generate the 'NG Altro' sheet using pre-fetched data.
    Dynamically creates columns based on unique `extra_data` values in 'Altro' defects.
    Returns True if any rows are written (excluding the header).
    """
    rows_written = 0
    id_moduli = data.get("id_moduli", [])
    objects = data.get("objects", [])
    productions = data.get("productions", [])
    stations = data.get("stations", [])
    production_lines = data.get("production_lines", [])
    object_defects = data.get("object_defects", [])

    objects_by_modulo = {obj["id_modulo"]: obj for obj in objects}
    stations_by_id = {station["id"]: station for station in stations}
    lines_by_id = {line["id"]: line for line in production_lines}

    # Step 1: gather unique Altro "extra_data" values
    unique_altro_descriptions = sorted({
        d.get("extra_data", "").strip()
        for d in object_defects
        if d.get("category") == "Altro" and d.get("extra_data")
    })

    # Step 2: Build the header
    header = [
        "Linea", "Stazione", "Stringatrice", "ID Modulo",
        "Data Ingresso", "Data Uscita", "Esito", "Tempo Ciclo"
    ] + [f"{desc}" for desc in unique_altro_descriptions]
    ws.append(header)

    for id_modulo in id_moduli:
        obj = objects_by_modulo.get(id_modulo)
        if not obj:
            continue

        object_id = obj["id"]
        prods = [p for p in productions if p.get("object_id") == object_id]
        if not prods:
            continue

        latest_prod = max(prods, key=lambda p: p.get("start_time") or 0)
        production_id = latest_prod["id"]
        start_time = latest_prod.get("start_time")
        end_time = latest_prod.get("end_time")
        esito = map_esito(latest_prod.get("esito"))
        cycle_time = str(latest_prod.get("cycle_time") or "")

        station = stations_by_id.get(latest_prod.get("station_id"))
        station_name = station.get("display_name", "Unknown") if station else "Unknown"

        line_name = "Unknown"
        if station and station.get("line_id"):
            line = lines_by_id.get(station["line_id"])
            if line:
                line_name = line.get("display_name", "Unknown")

        last_station_name = "N/A"
        last_station_id = latest_prod.get("last_station_id")
        if last_station_id:
            last_station = stations_by_id.get(last_station_id)
            if last_station:
                last_station_name = last_station.get("display_name", "N/A")

        # Get all "Altro" category defects for this production
        prod_defects = [
            d for d in object_defects
            if d.get("production_id") == production_id and d.get("category") == "Altro"
        ]
        if not prod_defects:
            continue

        altro_found = {d.get("extra_data", "").strip() for d in prod_defects}

        # Build NG flags for each Altro column
        altro_cols = ["NG" if desc in altro_found else "" for desc in unique_altro_descriptions]

        row = [
            line_name,
            station_name,
            last_station_name,
            id_modulo,
            start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else "",
            end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else "",
            esito,
            cycle_time
        ] + altro_cols

        ws.append(row)
        rows_written += 1

    if rows_written > 0:
        autofit_columns(ws, align_center_for=set(header))
        return True
    else:
        return False

# --- Mapping sheet names to functions ---
SHEET_FUNCTIONS = {
    "Metadata": metadata_sheet,
    "Risolutivo": risolutivo_sheet,
    "NG Generali": ng_generali_sheet,
    "NG Saldature": ng_saldature_sheet,
    "NG Disall. Ribbon": ng_disall_ribbon_sheet,
    "NG Disall. Stringa": ng_disall_stringa_sheet,
    "NG Mancanza Ribbon": ng_mancanza_ribbon_sheet,
    "NG Macchie ECA": ng_macchie_eca_sheet,
    "NG Celle Rotte": ng_celle_rotte_sheet,
    "NG Lunghezza String Ribbon": ng_lunghezza_string_ribbon_sheet,
    "NG Altro": ng_altro_sheet,
}

def map_esito(value: Optional[int]) -> str:
    if value == 1:
        return "OK"
    elif value == 2:
        return "In Produzione"
    elif value == 6:
        return "NG"
    return "N/A"


# ---------------- MYSQL ----------------
DB_SCHEMA = """
Il database contiene le seguenti tabelle:

1 `objects`
- id (PK)
- id_modulo (VARCHAR, UNIQUE)
- creator_station_id (FK to stations.id)
- created_at (DATETIME)

2 `stations`
- id (PK)
- line_id (FK to production_lines.id)
- name (VARCHAR)
- display_name (VARCHAR)
- type (ENUM: 'creator', 'qc', 'rework', 'other')
- config (JSON)
- created_at (DATETIME)

3 `production_lines`
- id (PK)
- name (VARCHAR)
- display_name (VARCHAR)
- description (TEXT)

4 `productions`
- id (PK)
- object_id (FK to objects.id)
- station_id (FK to stations.id)
- start_time (DATETIME)
- end_time (DATETIME)
- esito (INT) -- 1 = OK, 6 = KO, 2 = In Progress ( No Esito )
- operator_id (VARCHAR)
- cycle_time (TIME) -- calcolato come differenza tra end_time e start_time
- last_station_id (FK to stations.id, NULLABLE)

5 `defects`
- id (PK)
- category (ENUM: 'Generali', 'Saldatura', 'Disallineamento', 'Mancanza Ribbon', 'Macchie ECA', 'Celle Rotte', 'Lunghezza String Ribbon', 'Altro')

6 `object_defects`
- id (PK)
- production_id (FK to productions.id)
- defect_id (FK to defects.id)
- defect_type (VARCHAR, NULLABLE) -- usato solo per i "Generali"
- i_ribbon (INT, NULLABLE)
- stringa (INT, NULLABLE)
- ribbon_lato (ENUM: 'F', 'M', 'B', NULLABLE)
- s_ribbon (INT, NULLABLE)
- extra_data (VARCHAR, NULLABLE)

7 `station_defects`
- station_id (FK to stations.id)
- defect_id (FK to defects.id)
(Chiave primaria composta: station_id + defect_id)

"""""

async def insert_initial_production_data(data, station_name, connection):
    """
    Inserts a production record using data available at cycle start.
    It sets the end_time to NULL and uses esito = 2 (in progress).
    If a production record for the same object_id and station (with esito = 2 and end_time IS NULL)
    is already present, that record is returned instead of inserting a new row.
    """
    try:
        with connection.cursor() as cursor:
            id_modulo = data.get("Id_Modulo")
            # Determine line from the 'Linea_in_Lavorazione' list.
            linea_index = data.get("Linea_in_Lavorazione", [False] * 5).index(True) + 1
            actual_line = f"Linea{linea_index}"

            # Get line_id.
            cursor.execute("SELECT id FROM production_lines WHERE name = %s", (actual_line,))
            line_row = cursor.fetchone()
            if not line_row:
                raise ValueError(f"{actual_line} not found in production_lines")
            line_id = line_row["id"]

            # Get station id using station_name and line_id.
            cursor.execute("SELECT id FROM stations WHERE name = %s AND line_id = %s", (station_name, line_id))
            station_row = cursor.fetchone()
            if not station_row:
                raise ValueError(f"Station '{station_name}' not found for {actual_line}")
            real_station_id = station_row["id"]

            # Insert into objects table.
            sql_insert_object = """
                INSERT INTO objects (id_modulo, creator_station_id)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE id_modulo = id_modulo
            """
            cursor.execute(sql_insert_object, (id_modulo, real_station_id))

            # Get object_id.
            cursor.execute("SELECT id FROM objects WHERE id_modulo = %s", (id_modulo,))
            object_id = cursor.fetchone()["id"]

            # ☆ Check for existing partial production record:
            cursor.execute("""
                SELECT id FROM productions 
                WHERE object_id = %s 
                  AND station_id = %s 
                  AND esito = 2 
                  AND end_time IS NULL
                ORDER BY start_time DESC
                LIMIT 1
            """, (object_id, real_station_id))
            existing_prod = cursor.fetchone()
            if existing_prod:
                production_id = existing_prod["id"]
                connection.commit()
                logging.info(f"Production record already exists: ID {production_id} for object {object_id}")
                return production_id

            # Retrieve last_station_id from stringatrice if available.
            last_station_id = None
            str_flags = data.get("Lavorazione_Eseguita_Su_Stringatrice", [])
            if any(str_flags):
                stringatrice_index = str_flags.index(True) + 1
                stringatrice_name = f"Str{stringatrice_index}"
                cursor.execute(
                    "SELECT id FROM stations WHERE name = %s AND line_id = %s",
                    (stringatrice_name, line_id)
                )
                str_row = cursor.fetchone()
                if str_row:
                    last_station_id = str_row["id"]

            # Insert into productions table with esito = 2 (in progress) and no end_time.
            sql_productions = """
                INSERT INTO productions (
                    object_id, station_id, start_time, end_time, esito, operator_id, last_station_id
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(sql_productions, (
                object_id,
                real_station_id,
                data.get("DataInizio"),  # starting timestamp
                None,                     # end_time left as NULL
                2,                        # esito 2 means "in progress"
                data.get("Id_Utente"),
                last_station_id
            ))
            production_id = cursor.lastrowid

            connection.commit()
            logging.info(f"Initial production inserted: ID {production_id} for object {object_id}")
            return production_id

    except Exception as e:
        connection.rollback()
        logging.error(f"Error inserting initial production data: {e}")
        return None

async def update_production_final(production_id, data, station_name, connection):
    """
    Always update end_time. Update esito only if the current value is 2.
    """
    try:
        with connection.cursor() as cursor:
            # Step 1: Read current esito
            cursor.execute("SELECT esito FROM productions WHERE id = %s", (production_id,))
            row = cursor.fetchone()
            if not row:
                logging.warning(f"No production found with ID {production_id}")
                return False

            current_esito = row["esito"]
            final_esito = 6 if data.get("Compilato_Su_Ipad_Scarto_Presente") else 1
            end_time = data.get("DataFine")

            # Step 2: Conditional update
            if current_esito == 2:
                sql_update = """
                    UPDATE productions 
                    SET end_time = %s, esito = %s 
                    WHERE id = %s
                """
                cursor.execute(sql_update, (end_time, final_esito, production_id))
                logging.info(f"✅ Updated end_time + esito ({final_esito}) for production {production_id}")
            else:
                sql_update = """
                    UPDATE productions 
                    SET end_time = %s 
                    WHERE id = %s
                """
                cursor.execute(sql_update, (end_time, production_id))
                logging.info(f"ℹ️ Updated only end_time for production {production_id} (esito was already {current_esito})")

            connection.commit()
            return True

    except Exception as e:
        connection.rollback()
        logging.error(f"Error updating production {production_id}: {e}")
        return False

async def insert_defects(data, production_id, channel_id, line_name, cursor):
    # 1. Get defects mapping from DB.
    cursor.execute("SELECT id, category FROM defects")
    cat_map = {row["category"]: row["id"] for row in cursor.fetchall()}

    # 2. Load the issues from temporary storage using the proper line name.
    issues = get_latest_issues(line_name, channel_id)
    data["issues"] = issues  # Inject into data if needed later.

    # 3. For each issue path, parse and insert a row.
    for path in issues:
        category = detect_category(path)  # e.g. "Generali"
        defect_id = cat_map.get(category, cat_map["Altro"])  # fallback if unknown
        parsed = parse_issue_path(path, category)
        sql = """
            INSERT INTO object_defects (
                production_id, defect_id, defect_type, stringa, s_ribbon, i_ribbon, ribbon_lato, extra_data
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        cursor.execute(sql, (
            production_id,
            defect_id,
            parsed["defect_type"],
            parsed["stringa"],
            parsed["s_ribbon"],
            parsed["i_ribbon"],
            parsed["ribbon_lato"],
            parsed['extra_data']
        ))

async def update_esito(esito: int, production_id: int, cursor):
    """
    Update the 'esito' field in the productions table for the given production ID.
    """
    try:
        sql_update = """
            UPDATE productions 
            SET esito = %s 
            WHERE id = %s
        """
        cursor.execute(sql_update, (esito, production_id))
        return True
    except Exception as e:
        if mysql_connection:
            mysql_connection.rollback()
        logging.error(f"❌ Error updating esito for production_id={production_id}: {e}")
        return False
    
def detect_category(path: str) -> str:
    parts = path.split(".")
    if len(parts) < 5:
        return "Altro"
    category_raw = parts[4]
    if ":" in category_raw:
        category_raw = category_raw.split(":")[0].strip()
    return category_raw

def parse_issue_path(path: str, category: str):
    """
    Returns a dict with fields:
      {
        "defect_type": ...   # For Generali
        "stringa": ...       # For Saldatura, Disallineamento, Macchie ECA, etc.
        "s_ribbon": ...      # For Saldatura (Pin)
        "i_ribbon": ...      # For Disallineamento/Mancanza Ribbon
        "ribbon_lato": ...   # Possibly 'F','M','B'
      }
    """
    res: Dict[str, Optional[Union[str, int]]] = {
        "defect_type": None,
        "stringa": None,
        "s_ribbon": None,
        "i_ribbon": None,
        "ribbon_lato": None,
        "extra_data": None
    }

    # Split
    parts = path.split(".")

    if category == "Generali":
        # last part might be the actual defect, e.g. "Bad Soldering"
        # path might be: "...Generali.Bad Soldering"
        if len(parts) >= 5:
            res["defect_type"] = parts[-1]  # e.g. "Bad Soldering"
        return res

    elif category == "Saldatura":
        # path e.g. "Dati.Esito.Esito_Scarto.Difetti.Saldatura.Stringa[2].Pin[5].M"
        # Let's parse out 'Stringa[2]' => stringa=2
        # 'Pin[5]' => s_ribbon=5
        # 'M' => ribbon_lato='M'
        # That might be parts[4], parts[5], parts[6]
        # e.g. parts[4]="Stringa[2]", parts[5]="Pin[5]", parts[6]="M"
        str_match = re.search(r"Stringa\[(\d+)\]", path)
        pin_match = re.search(r"Pin\[(\d+)\]", path)
        side_match = re.search(r"\ - (F|M|B)$", path)  # the last segment

        if str_match:
            res["stringa"] = int(str_match.group(1))
        if pin_match:
            res["s_ribbon"] = int(pin_match.group(1))
        if side_match:
            res["ribbon_lato"] = side_match.group(1)

    elif category == "Disallineamento":
        # Could be: "...Disallineamento.Stringa[3]" or "...Disallineamento.Ribbon[5].F"
        str_match = re.search(r"Stringa\[(\d+)\]", path)
        if str_match:
            res["stringa"] = int(str_match.group(1))
        else:
            # else check Ribbon
            # e.g. "Disallineamento.Ribbon[5].M"
            rib_match = re.search(r"Ribbon\[(\d+)\]", path)
            side_match = re.search(r"\ - (F|M|B)$", path)
            if rib_match:
                res["i_ribbon"] = int(rib_match.group(1))
            if side_match:
                res["ribbon_lato"] = side_match.group(1)

    elif category == "Mancanza Ribbon":
        # e.g. "Mancanza Ribbon.Ribbon[2].B"
        # i_ribbon=2, ribbon_lato='B'
        rib_match = re.search(r"Ribbon\[(\d+)\]", path)
        side_match = re.search(r"\ - (F|M|B)$", path)
        if rib_match:
            res["i_ribbon"] = int(rib_match.group(1))
        if side_match:
            res["ribbon_lato"] = side_match.group(1)

    elif category == "Macchie ECA":
        # e.g. "Macchie ECA.Stringa[4]"
        str_match = re.search(r"Stringa\[(\d+)\]", path)
        if str_match:
            res["stringa"] = int(str_match.group(1))

    elif category == "Celle Rotte":
        # e.g. "Celle Rotte.Stringa[6]"
        str_match = re.search(r"Stringa\[(\d+)\]", path)
        if str_match:
            res["stringa"] = int(str_match.group(1))

    elif category == "Lunghezza String Ribbon":
        # e.g. "Lunghezza String Ribbon.Stringa[2]"
        str_match = re.search(r"Stringa\[(\d+)\]", path)
        if str_match:
            res["stringa"] = int(str_match.group(1))

    elif category == "Altro":
        # Example: "Dati.Esito.Esito_Scarto.Difetti.Altro: Macchia sulla cella"
        print('ALTRO Path: %s' % path)
        if ":" in path:
            res["extra_data"] = path.split(":", 1)[1].strip()


    return res


# ---------------- CONFIG PARSING ----------------
def load_station_configs(file_path):
    config = configparser.ConfigParser()
    config.read(file_path)

    line_configs = {}

    current_line = None
    for section in config.sections():
        if section.upper().startswith("LINEA"):
            current_line = section
            line_configs[current_line] = {
                "PLC": {},
                "PC": {},
                "stations": []
            }

        elif section.upper() == "PLC" and current_line:
            line_configs[current_line]["PLC"]["IP"] = config.get(section, "IP")
            line_configs[current_line]["PLC"]["SLOT"] = config.getint(section, "SLOT")

        elif section.upper() == "PC" and current_line:
            line_configs[current_line]["PC"]["IP"] = config.get(section, "IP")
            line_configs[current_line]["PC"]["PORT"] = config.getint(section, "PORT")

        elif section.startswith("M") and current_line:
            line_configs[current_line]["stations"].append(section)

    return line_configs


# ---------------- BACKGROUND TASK ----------------
def make_status_callback(full_station_id: str):
    async def callback(status):
        try:
            line_name, channel_id = full_station_id.split(".")
            await broadcast(line_name, channel_id, {"plc_status": status})
        except Exception as e:
            logging.error(f"❌ Failed to send PLC status for {full_station_id}: {e}")
    return callback

async def background_task(plc_connection: PLCConnection, full_station_id: str):
    print(f"[{full_station_id}] Starting background task.")
    prev_trigger = False

    line_name, channel_id = full_station_id.split(".")

    while True:
        try:
            # Ensure connection is alive or try reconnect
            if not plc_connection.connected or not plc_connection.is_connected():
                print(f"⚠️ PLC disconnected for {full_station_id}, attempting reconnect...")
                await asyncio.to_thread(plc_connection.reconnect, retries=3, delay=5)
                await asyncio.sleep(10)
                continue  # Retry after delay

            paths = get_channel_config(line_name, channel_id)
            if not paths:
                logging.error(f"❌ Invalid line/channel: {line_name}.{channel_id}")
                await asyncio.sleep(1)
                continue  # Skip this cycle if config not found

            trigger_conf = paths["trigger"]
            trigger_value = await asyncio.to_thread(
                plc_connection.read_bool,
                trigger_conf["db"], trigger_conf["byte"], trigger_conf["bit"]
            )


            if trigger_value is None:
                raise Exception("Trigger read returned None")

            if trigger_value != prev_trigger:
                prev_trigger = trigger_value
                await on_trigger_change(plc_connection, line_name, channel_id, None, trigger_value, None)

            # Outcome check
            paths = get_channel_config(line_name, channel_id)
            if not paths:
                logging.error(f"❌ Missing config for {line_name}.{channel_id}")
                await asyncio.sleep(1)
                continue  # Or return / skip, depending on context

            # Now you're safe to use it:
            fb_conf = paths["fine_buona"]
            fs_conf = paths["fine_scarto"]
            fine_buona = await asyncio.to_thread(plc_connection.read_bool, fb_conf["db"], fb_conf["byte"], fb_conf["bit"])
            fine_scarto = await asyncio.to_thread(plc_connection.read_bool, fs_conf["db"], fs_conf["byte"], fs_conf["bit"])


            if fine_buona is None or fine_scarto is None:
                raise Exception("Outcome read returned None")

            if (fine_buona or fine_scarto) and not passato_flags[full_station_id]:
                data_inizio = trigger_timestamps.get(full_station_id)
                result = await read_data(plc_connection, line_name, channel_id,
                                         richiesta_ok=fine_buona,
                                         richiesta_ko=fine_scarto,
                                         #richiesta_ok=False,
                                         #richiesta_ko=True,
                                         data_inizio=data_inizio)
                if result:
                    passato_flags[full_station_id] = True
                    production_id = incomplete_productions.get(full_station_id)
                    if production_id:
                        # Update the initial production record with final data.
                        await update_production_final(production_id, result, channel_id, mysql_connection)
                        incomplete_productions.pop(full_station_id)
                    else:
                        logging.warning(f"⚠️ No initial production record found for {full_station_id}; skipping update.")
                    remove_temp_issues(line_name, channel_id, result.get("Id_Modulo"))

            await asyncio.sleep(1)

        except Exception as e:
            logging.error(f"[{full_station_id}] 🔴 Error in background task: {str(e)}")
            await asyncio.to_thread(plc_connection.reconnect, retries=3, delay=5)
            await asyncio.sleep(5)


# ---------------- WEB SOCKET ----------------
@app.websocket("/ws/summary/{line_name}")
async def websocket_summary(websocket: WebSocket, line_name: str):
    await websocket.accept()
    key = f"{line_name}.summary"
    print(f"📊 Dashboard summary client connected for {line_name}")

    subscriptions.setdefault(key, set()).add(websocket)

    try:
        while True:
            await websocket.receive_text()  # Or just keep alive
    except WebSocketDisconnect:
        print(f"❌ Dashboard summary client for {line_name} disconnected")
        subscriptions[key].remove(websocket)

@app.websocket("/ws/{line_name}/{channel_id}")
async def websocket_endpoint(websocket: WebSocket, line_name: str, channel_id: str):
    full_id = f"{line_name}.{channel_id}"

    if get_channel_config(line_name, channel_id) is None:
        await websocket.close()
        print(f"❌ Invalid channel config for {full_id}")
        return

    await websocket.accept()
    await websocket.send_json({"handshake": True})
    print(f"📲 Client subscribed to {full_id}")

    subscriptions.setdefault(full_id, set()).add(websocket)
    plc_connection = plc_connections.get(full_id)

    if not plc_connection:
        print(f"❌ No PLC connection found for {full_id}.")
        await websocket.close()
        return

    # Check connection status before sending initial state
    if not plc_connection.connected or not plc_connection.is_connected():
        print(f"⚠️ PLC for {full_id} is disconnected. Attempting reconnect for WebSocket...")
        if not plc_connection.reconnect(retries=3, delay=5):
            print(f"❌ Failed to reconnect PLC for {full_id}. Closing socket.")
            await websocket.close()
            return
        else:
            print(f"✅ PLC reconnected for {full_id}!")

    await send_initial_state(websocket, channel_id, plc_connection, line_name)

    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        print(f"⚠️ Client disconnected from {full_id}")
    finally:
        subscriptions[full_id].remove(websocket)


# ---------------- HELPER ----------------
def generate_unique_filename(base_path, base_name, extension):
    i = 1
    full_path = os.path.join(base_path, f"{base_name}{extension}")
    while os.path.exists(full_path):
        full_path = os.path.join(base_path, f"{base_name}_{i}{extension}")
        i += 1
    return full_path


# ---------------- ROUTES ----------------
@app.get("/api/plc_status")
async def plc_status():
    statuses = {}

    for full_id, plc_conn in plc_connections.items():
        try:
            line_name, channel_id = full_id.split(".")
        except ValueError:
            continue  # Skip malformed keys, just in case

        if line_name not in statuses:
            statuses[line_name] = {}

        statuses[line_name][channel_id] = "CONNECTED" if plc_conn.connected else "DISCONNECTED"

    return statuses

@app.post("/api/upload_images")
async def upload_images(
    object_id: str = Form(...),
    images: List[UploadFile] = File(...),
    defects: List[str] = Form(...)
):
    folder_path = f"C:/IX-Monitor/images_submitted/{object_id}"
    os.makedirs(folder_path, exist_ok=True)

    for i, image_file in enumerate(images):
        defect_name = defects[i].replace(" ", "_").replace(":", "_")
        filename = image_file.filename or f"image_{i}.jpg"
        ext = os.path.splitext(filename)[-1]
        save_path = generate_unique_filename(folder_path, defect_name, ext)

        with open(save_path, "wb") as f:
            content = await image_file.read()
            f.write(content)

    return {"status": "ok", "saved": len(images)}

@app.post("/api/set_issues")
async def set_issues(request: Request):
    data = await request.json()
    line_name = data.get("line_name")
    channel_id = data.get("channel_id")
    object_id = data.get("object_id")
    issues = data.get("issues", [])

    if not line_name or not channel_id or not object_id or not issues:
        return JSONResponse(status_code=400, content={"error": "Missing data"})

    full_id = f"{line_name}.{channel_id}"

    # Load existing data
    existing_data = load_temp_data()

    # Append new data
    existing_data.append({
        "line_name": line_name,
        "channel_id": channel_id,
        "object_id": object_id,
        "issues": issues
    })

    # Save updated data
    save_temp_data(existing_data)

    plc_connection = plc_connections.get(full_id)
    if not plc_connection:
        return JSONResponse(status_code=404, content={"error": f"No PLC connection for {full_id}."})

    paths = get_channel_config(line_name, channel_id)
    if not paths:
        return JSONResponse(status_code=404, content={"error": "Channel mapping not found"})

    target = paths.get("esito_scarto_compilato")
    if not target:
        return JSONResponse(status_code=404, content={"error": "esito_scarto_compilato not found in mapping"})
    
    # 💾 Immediately insert defects if production is already tracked
    production_id = incomplete_productions.get(full_id)
    if production_id:
        try:
            assert mysql_connection is not None
            cursor = mysql_connection.cursor()
            result = {
                "Id_Modulo": object_id,
                "Compilato_Su_Ipad_Scarto_Presente": True,
                "issues": issues
            }
            await insert_defects(result, production_id, channel_id, line_name, cursor=cursor)
            await update_esito(6, production_id, cursor=cursor)
            mysql_connection.commit()
            cursor.close()
            print(f"✅ Defects inserted immediately for {object_id} (prod_id: {production_id})")
        except Exception as e:
            assert mysql_connection is not None
            mysql_connection.rollback()
            logging.error(f"❌ Error inserting defects early for {full_id}: {e}")


    await asyncio.to_thread(plc_connection.write_bool, target["db"], target["byte"], target["bit"], True)

    return {"status": "ok"}

@app.get("/api/get_issues")
async def get_selected_issues(line_name: str, channel_id: str, object_id: str):
    if not line_name or not channel_id or not object_id:
        return JSONResponse(status_code=400, content={"error": "Missing line_name, channel_id, or object_id"})

    # Load issues from temp storage
    temp_data = load_temp_data()

    # Filter issues based on line_name + channel_id + object_id
    matching_entry = next(
        (
            entry for entry in temp_data
            if entry.get("line_name") == line_name and
               entry.get("channel_id") == channel_id and
               entry.get("object_id") == object_id
        ),
        None
    )

    if not matching_entry:
        return {"selected_issues": []}  # Return empty if not found

    return {"selected_issues": matching_entry["issues"]}

@app.post("/api/set_outcome")
async def set_outcome(request: Request):
    data = await request.json()
    line_name = data.get("line_name")
    channel_id = data.get("channel_id")
    object_id = data.get("object_id")
    outcome = data.get("outcome")  # "buona" or "scarto"

    if not line_name or not channel_id or outcome not in ["buona", "scarto"]:
        return JSONResponse(status_code=400, content={"error": "Invalid data"})

    # Get channel config
    config = get_channel_config(line_name, channel_id)
    if not config:
        return JSONResponse(status_code=404, content={"error": "Channel not found"})

    plc_connection = plc_connections.get(f"{line_name}.{channel_id}")
    if not plc_connection:
        return JSONResponse(status_code=404, content={"error": "PLC connection not found"})

    # Read current object_id from PLC
    read_conf = config["id_modulo"]
    current_object_id = await asyncio.to_thread(
        plc_connection.read_string,
        read_conf["db"], read_conf["byte"], read_conf["length"]
    )

    if str(current_object_id) != str(object_id):
        return JSONResponse(status_code=409, content={"error": "Stale object, already processed or expired."})

    # Optional: Write outcome back to PLC if needed (currently commented out)
    # outcome_conf = config["fine_buona"] if outcome == "buona" else config["fine_scarto"]
    # await asyncio.to_thread(
    #     plc_connection.write_bool,
    #     outcome_conf["db"], outcome_conf["byte"], outcome_conf["bit"], True
    # )

    print(f"✅ Outcome '{outcome.upper()}' written for object {object_id} on {line_name}.{channel_id}")
    await broadcast(line_name, channel_id, {
        "trigger": None,
        "objectId": object_id,
        "outcome": outcome
    })

    return {"status": "ok"}

@app.get("/api/issues/{line_name}/{channel_id}")
async def get_issue_tree(
    line_name: str,
    channel_id: str,
    path: str = Query("Dati.Esito.Esito_Scarto.Difetti")
):
    # Validate channel
    config = get_channel_config(line_name, channel_id)
    if not config:
        return JSONResponse(status_code=404, content={"error": "Invalid line or channel"})

    # Traverse ISSUE_TREE using the dot-separated path
    current_node = ISSUE_TREE
    if path:
        for part in path.split("."):
            current_node = current_node.get(part)
            if current_node is None:
                return JSONResponse(status_code=404, content={"error": f"Path '{path}' not found"})

    items = []
    for name, child in current_node.items():
        item_type = "folder" if child else "leaf"
        items.append({"name": name, "type": item_type})

    return {"items": items}

@app.get("/api/overlay_config")
async def get_overlay_config(
    path: str,
    line_name: str = Query(...),
    station: str = Query(...)
):
    config_file = f"C:/IX-Monitor/images/{line_name}/{station}/overlay_config.json"
    print(f"🔍 Requested overlay config for path: '{path}' (line: {line_name}, station: {station})")
    print(f"📄 Looking for config file at: {config_file}")

    if not os.path.exists(config_file):
        print(f"❌ Config file not found.")
        return JSONResponse(status_code=417, content={"error": f"Overlay config not found for {line_name}/{station}"})

    try:
        with open(config_file, "r") as f:
            all_configs = json.load(f)
            print(f"✅ Loaded overlay_config.json with keys: {list(all_configs.keys())}")
    except json.JSONDecodeError as e:
        print(f"❌ JSON decode error: {e}")
        all_configs = {}

    for image_name, config in all_configs.items():
        config_path = config.get("path", "")
        print(f"➡️ Checking config: image = '{image_name}', path = '{config_path}'")

        if config_path.lower() == path.lower():
            image_url = f"http://192.168.0.10:8000/images/{line_name}/{station}/{image_name}"
            print(f"✅ MATCH FOUND! Returning image: {image_url}")
            return {
                "image_url": image_url,
                "rectangles": config.get("rectangles", [])
            }

    print(f"⚠️ No matching path found. Returning fallback with empty image URL.")
    return {
        "image_url": "",
        "rectangles": []
    }

@app.post("/api/update_overlay_config")
async def update_overlay_config(request: Request):
    data = await request.json()
    path = data.get("path")
    new_rectangles = data.get("rectangles")
    line_name = data.get("line_name")
    station = data.get("station")

    if not path or not new_rectangles or not line_name or not station:
        return JSONResponse(status_code=400, content={"error": "Missing path, rectangles, line_name, or station"})

    config_file = f"C:/IX-Monitor/images/{line_name}/{station}/overlay_config.json"

    if not os.path.exists(config_file):
        return JSONResponse(status_code=404, content={"error": f"Config file not found for {line_name}/{station}"})

    with open(config_file, "r") as f:
        config = json.load(f)

    image_to_update = None
    for image_name, entry in config.items():
        if entry.get("path") == path:
            image_to_update = image_name
            break

    if not image_to_update:
        return JSONResponse(status_code=404, content={"error": "Path not found in config"})

    config[image_to_update]["rectangles"] = new_rectangles

    with open(config_file, "w") as f:
        json.dump(config, f, indent=4)

    return {"status": "updated", "image": image_to_update}

@app.get("/api/available_overlay_paths")
async def available_overlay_paths(
    line_name: str = Query(...),
    station: str = Query(...)
):
    config_file = f"C:/IX-Monitor/images/{line_name}/{station}/overlay_config.json"

    if not os.path.exists(config_file):
        return JSONResponse(status_code=404, content={"error": f"Config file not found for {line_name}/{station}"})

    with open(config_file, "r") as f:
        all_configs = json.load(f)

    result = []
    for image_name, config in all_configs.items():
        path = config.get("path")
        if path:
            result.append({
                "image": image_name,
                "path": path
            })
    return result

@app.get("/api/productions_summary")
async def productions_summary(
    date: Optional[str] = Query(default=None),
    from_date: Optional[str] = Query(default=None, alias="from"),
    to_date: Optional[str] = Query(default=None, alias="to"),
    line_name: Optional[str] = Query(default=None),
    turno: Optional[int] = Query(default=None),
    start_time: Optional[str] = Query(default=None),
    end_time: Optional[str] = Query(default=None),
):
    global mysql_connection
    try:
        assert mysql_connection is not None
        with mysql_connection.cursor() as cursor:

            # Build base WHERE clause for production filtering on end_time (new schema)
            params = []
            where_clause = ""
            if from_date and to_date:
                where_clause = "WHERE DATE(p.end_time) BETWEEN %s AND %s"
                params.extend([from_date, to_date])
            elif date:
                if turno and turno == 3:
                    # For turno 3, use a datetime range: from 22:00 on the selected date to 05:59:59 on the next day.
                    start_dt = f"{date} 22:00:00"
                    end_dt = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d 05:59:59")
                    where_clause = "WHERE p.end_time BETWEEN %s AND %s"
                    params.extend([start_dt, end_dt])
                else:
                    where_clause = "WHERE DATE(p.end_time) = %s"
                    params.append(date)
            else:
                return JSONResponse(status_code=400, content={"error": "Missing 'date' or 'from' and 'to'"})

            if turno:
                turno_times = {
                    1: ("06:00:00", "13:59:59"),
                    2: ("14:00:00", "21:59:59"),
                    3: ("22:00:00", "05:59:59"),
                }
                if turno not in turno_times:
                    return JSONResponse(
                        status_code=400,
                        content={"error": "Invalid turno number (must be 1, 2, or 3)"}
                    )

                turno_start, turno_end = turno_times[turno]

                if turno == 3:
                    if date:
                        shift_day = datetime.strptime(date, "%Y-%m-%d")
                        next_day = shift_day + timedelta(days=1)
                        where_clause += """
                            AND (
                                (DATE(p.end_time) = %s AND TIME(p.end_time) >= '22:00:00')
                                OR
                                (DATE(p.end_time) = %s AND TIME(p.end_time) <= '05:59:59')
                            )
                        """
                        params.extend([shift_day.strftime("%Y-%m-%d"), next_day.strftime("%Y-%m-%d")])
                    elif from_date and to_date:
                        where_clause += """
                            AND (
                                TIME(p.end_time) >= '22:00:00'
                                OR TIME(p.end_time) <= '05:59:59'
                            )
                        """
                    else:
                        return JSONResponse(
                            status_code=400,
                            content={"error": "Missing 'date' or 'from' and 'to'"}
                        )
                else:
                    where_clause += " AND TIME(p.end_time) BETWEEN %s AND %s"
                    params.extend([turno_start, turno_end])

            # Optional start_time and end_time overrides
            if start_time and end_time:
                try:
                    # Convert to proper datetime strings if needed
                    _ = datetime.fromisoformat(start_time)
                    _ = datetime.fromisoformat(end_time)
                    where_clause += " AND p.end_time BETWEEN %s AND %s"
                    params.extend([start_time, end_time])
                except ValueError:
                    return JSONResponse(
                        status_code=400,
                        content={"error": "start_time and end_time must be ISO 8601 formatted strings"}
                    )

            # Filter by production line if provided.
            # In the new schema, we join production_lines (alias pl) via stations.
            if line_name:
                try:
                    where_clause += " AND pl.name = %s"
                    params.append(line_name)
                except ValueError:
                    return JSONResponse(status_code=400, content={"error": "Invalid line_name format"})

            # --- Summary per station ---
            # Join productions with stations (and production_lines for filtering) to use new schema columns.
            query = f"""
                SELECT
                    s.name AS station_name,
                    s.display_name AS station_display,
                    SUM(CASE WHEN p.esito = 1 THEN 1 ELSE 0 END) AS good_count,
                    SUM(CASE WHEN p.esito = 6 THEN 1 ELSE 0 END) AS bad_count,
                    SEC_TO_TIME(AVG(TIME_TO_SEC(p.cycle_time))) AS avg_cycle_time
                FROM productions p
                JOIN stations s ON p.station_id = s.id
                LEFT JOIN production_lines pl ON s.line_id = pl.id
                {where_clause}
                GROUP BY s.name, s.display_name
            """
            cursor.execute(query, tuple(params))
            stations = {}
            for row in cursor.fetchall():
                name = row['station_name']
                stations[name] = {
                    "display": row['station_display'],  # M308 - QG2 di M306
                    "good_count": int(row['good_count']),
                    "bad_count": int(row['bad_count']),
                    "avg_cycle_time": str(row['avg_cycle_time']),
                    "last_cycle_time": "00:00:00"
                }

            print('stations: ', stations)

            # Fill in missing stations (for visual consistency).
            # CHANNELS should be defined elsewhere in your code.
            all_station_names = [
                station for line, stations_list in CHANNELS.items()
                if line == line_name or line_name is None
                for station in stations_list
            ]
            for station in all_station_names:
                stations.setdefault(station, {
                    "display": station,  # fallback to code if display is missing
                    "good_count": 0,
                    "bad_count": 0,
                    "avg_cycle_time": "00:00:00",
                    "last_cycle_time": "00:00:00"
                })


            # --- Defect Summary ---
            # Helper function to fetch defects by category.
            def fetch_defect_summary(category_label, label):
                q = f"""
                    SELECT 
                        s.name AS station_code, 
                        d.category,
                        COUNT(DISTINCT CONCAT(od.production_id, '-', d.category)) AS unique_defect_count
                    FROM object_defects od
                    JOIN defects d ON od.defect_id = d.id
                    JOIN productions p ON p.id = od.production_id
                    JOIN stations s ON p.station_id = s.id
                    LEFT JOIN production_lines pl ON s.line_id = pl.id
                    {where_clause} AND p.esito = 6
                    GROUP BY s.name, d.category
                """
                
                cursor.execute(q, tuple(params))  # ✅ Use params only

                for row in cursor.fetchall():
                    station_code = row['station_code']
                    category = row['category']
                    count = int(row['unique_defect_count'])
                    if station_code in stations:
                        stations[station_code].setdefault("defects", {})[category] = count

            fetch_defect_summary("Mancanza Ribbon", "Mancanza Ribbon")
            fetch_defect_summary("Saldatura", "Saldatura")
            fetch_defect_summary("Disallineamento", "Disallineamento")
            fetch_defect_summary("Generali", "Generali")

            # Calculate "KO Generico" (generic KO) for each station.
            for station, data in stations.items():
                bad_count_val = int(data["bad_count"])
                defects = data.get("defects", {})
                total_defects = sum(defects.values())
                generic = bad_count_val - total_defects
                if generic > 0:
                    stations[station].setdefault("defects", {})["Generico"] = generic

            # --- Last cycle time details (always fetch latest per station) ---
            query_last = f"""
                SELECT s.name as station, o.id_modulo, p.esito, p.cycle_time, p.start_time, p.end_time
                FROM productions p
                JOIN stations s ON p.station_id = s.id
                JOIN objects o ON p.object_id = o.id
                LEFT JOIN production_lines pl ON s.line_id = pl.id
                {where_clause}
                ORDER BY p.end_time DESC
            """
            # ✅ This is the correct version:
            cursor.execute(query_last, tuple(params))

            seen_stations = set()
            for row in cursor.fetchall():
                station = row['station']
                if station not in seen_stations and station in stations:
                    stations[station]["last_object"] = row["id_modulo"]
                    stations[station]["last_esito"] = row["esito"]
                    stations[station]["last_cycle_time"] = str(row["cycle_time"])
                    stations[station]["last_in_time"] = str(row["start_time"])
                    stations[station]["last_out_time"] = str(row["end_time"])
                    seen_stations.add(station)

            good_count_total = sum(s["good_count"] for s in stations.values())
            bad_count_total = sum(s["bad_count"] for s in stations.values())

            return {
                "good_count": good_count_total,
                "bad_count": bad_count_total,
                "stations": stations,
            }

    except Exception as e:
        logging.error(f"MySQL Error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

# ---------------- SEARCH ----------------
# Mapping from client order field names to DB columns.
column_map = {
    "ID Modulo": "o.id_modulo",
    "Esito": "p.esito",
    "Data": "p.end_time",
    "Operatore": "p.operator_id",
    "Linea": "pl.display_name",
    "Stazione": "s.name",
    "Tempo Ciclo": "p.cycle_time"
}

@app.post("/api/search")
async def search_results(request: Request):
    global mysql_connection
    try:
        payload = await request.json()
        print(payload)

        filters: List[Dict[str, str]] = payload.get("filters", [])
        order_by_input = payload.get("order_by", "Data")
        order_by = column_map.get(order_by_input, "p.end_time")
        order_direction: str = payload.get("order_direction", "DESC")
        limit: int = int(payload.get("limit", 1000))
        direction = "ASC" if order_direction.lower() == "crescente" else "DESC"

        # Initialize join and where clauses.
        join_clauses = []
        where_clauses = []
        params = []

        # If any filter is for defects, include the defects join.
        has_defect_filter = any(f.get("type") == "Difetto" for f in filters)
        if has_defect_filter:
            join_clauses.append("JOIN object_defects od ON p.id = od.production_id")
            join_clauses.append("JOIN defects d ON od.defect_id = d.id")

        # Process all filters.
        for f in filters:
            key, value = f.get("type"), f.get("value")
            if key == "ID Modulo":
                where_clauses.append("o.id_modulo LIKE %s")
                params.append(f"%{value}%")
            
            elif key == "Esito":
                if value == "OK":
                    where_clauses.append("p.esito = 1")
                elif value == "KO":
                    where_clauses.append("p.esito = 6")
                elif value == "In Produzione":
                    where_clauses.append("p.esito NOT IN (1, 6)")
            
            elif key == "Operatore":
                where_clauses.append("p.operator_id LIKE %s")
                params.append(f"%{value}%")
            
            elif key == "Linea":
                where_clauses.append("pl.display_name = %s")
                params.append(value)
            
            elif key == "Stazione":
                where_clauses.append("s.name = %s")
                params.append(value)
            
            elif key == "Turno":
                turno_times = {
                    "1": ("06:00:00", "13:59:59"),
                    "2": ("14:00:00", "21:59:59"),
                    "3": ("22:00:00", "05:59:59"),
                }
                if value in turno_times:
                    start_t, end_t = turno_times[value]
                    if value == "3":
                        where_clauses.append(
                            "(TIME(p.end_time) >= '22:00:00' OR TIME(p.end_time) <= '05:59:59')"
                        )
                    else:
                        where_clauses.append("TIME(p.end_time) BETWEEN %s AND %s")
                        params.extend([start_t, end_t])
            
            elif key == "Data":
                from_iso = f.get("start")
                to_iso = f.get("end")
                if from_iso and to_iso:  # Make sure they're not None
                    try:
                        from_dt = datetime.fromisoformat(from_iso)
                        to_dt = datetime.fromisoformat(to_iso)
                        where_clauses.append("p.end_time BETWEEN %s AND %s")
                        params.extend([from_dt, to_dt])
                    except Exception as e:
                        print(f"ISO datetime parsing error: {e}")
                else:
                    print("Missing 'start' or 'end' value in Data filter")

            elif key == "Stringatrice":
                stringatrice_map = {
                    "1": "Str1",
                    "2": "Str2",
                    "3": "Str3",
                    "4": "Str4",
                    "5": "Str5"
                }
                if value in stringatrice_map:
                    # Add a second join to stations table (aliased as ls for last_station)
                    join_clauses.append("LEFT JOIN stations ls ON p.last_station_id = ls.id")
                    where_clauses.append("ls.name = %s")
                    params.append(stringatrice_map[value])
            
            elif key == "Difetto" and value:
                # Split the composite defect filter by " > "
                parts = value.split(" > ")
                if not parts:
                    continue
                # Always filter by defect category.
                defect_category = parts[0]
                where_clauses.append("d.category = %s")
                params.append(defect_category)

                if defect_category == "Generali":
                    if len(parts) > 1 and parts[1].strip():
                        # For Generali, second part is defect_type.
                        where_clauses.append("od.defect_type = %s")
                        params.append(parts[1])
                elif defect_category == "Saldatura":
                    # Expected: Saldatura > Stringa[<num>] > Lato <letter> > Pin[<num>]
                    if len(parts) > 1 and parts[1].strip():
                        match = re.search(r'\[(\d+)\]', parts[1])
                        if match:
                            stringa_num = int(match.group(1))
                            where_clauses.append("od.stringa = %s")
                            params.append(stringa_num)
                    if len(parts) > 2 and parts[2].strip():
                        lato = parts[2].replace("Lato ", "").strip()
                        where_clauses.append("od.ribbon_lato = %s")
                        params.append(lato)
                    if len(parts) > 3 and parts[3].strip():
                        match = re.search(r'\[(\d+)\]', parts[3])
                        if match:
                            pin_num = int(match.group(1))
                            where_clauses.append("od.s_ribbon = %s")
                            params.append(pin_num)
                elif defect_category == "Disallineamento":
                    # Expected can be either:
                    #   "Disallineamento > Stringa > Stringa[<num>]"
                    #   or "Disallineamento > Ribbon > Lato <letter> > Ribbon[<num>]"
                    if len(parts) > 1 and parts[1].strip():
                        mode = parts[1]
                        if mode == "Stringa":
                            if len(parts) > 2 and parts[2].strip():
                                match = re.search(r'\[(\d+)\]', parts[2])
                                if match:
                                    stringa_num = int(match.group(1))
                                    where_clauses.append("od.stringa = %s")
                                    params.append(stringa_num)
                        elif mode == "Ribbon":
                            if len(parts) > 2 and parts[2].strip():
                                lato = parts[2].replace("Lato ", "").strip()
                                where_clauses.append("od.ribbon_lato = %s")
                                params.append(lato)
                            if len(parts) > 3 and parts[3].strip():
                                match = re.search(r'\[(\d+)\]', parts[3])
                                if match:
                                    ribbon_num = int(match.group(1))
                                    # For Disallineamento in Ribbon mode, use i_ribbon.
                                    where_clauses.append("od.i_ribbon = %s")
                                    params.append(ribbon_num)
                elif defect_category == "Mancanza Ribbon":
                    # Expected: Mancanza Ribbon > Lato <letter> > Ribbon[<num>]
                    if len(parts) > 1 and parts[1].strip():
                        lato = parts[1].replace("Lato ", "").strip()
                        where_clauses.append("od.ribbon_lato = %s")
                        params.append(lato)
                    if len(parts) > 2 and parts[2].strip():
                        match = re.search(r'\[(\d+)\]', parts[2])
                        if match:
                            ribbon_num = int(match.group(1))
                            # For Mancanza Ribbon use i_ribbon.
                            where_clauses.append("od.i_ribbon = %s")
                            params.append(ribbon_num)
                elif defect_category in ("Macchie ECA", "Celle Rotte", "Lunghezza String Ribbon"):
                    # Expected: e.g., "Macchie ECA > Stringa[<num>]"
                    if len(parts) > 1 and parts[1].strip():
                        match = re.search(r'\[(\d+)\]', parts[1])
                        if match:
                            stringa_num = int(match.group(1))
                            where_clauses.append("od.stringa = %s")
                            params.append(stringa_num)
                elif defect_category == "Altro":
                    if len(parts) > 1 and parts[1].strip():
                        where_clauses.append("od.extra_data LIKE %s")
                        params.append(f"%{parts[1]}%")

        # Build final JOIN and WHERE SQL parts.
        join_sql = " ".join(join_clauses)
        where_sql = " AND ".join(where_clauses)
        if where_sql:
            where_sql = "WHERE " + where_sql

        # Special ordering logic for NULLs in esito/cycle_time.
        if order_by in {"p.esito", "p.cycle_time"}:
            order_clause = f"ORDER BY ISNULL({order_by}), {order_by} {direction}"
        else:
            order_clause = f"ORDER BY {order_by} {direction}"

        select_fields = """
            o.id_modulo, 
            p.esito, 
            p.operator_id, 
            p.cycle_time, 
            p.start_time, 
            p.end_time,
            s.name AS station_name,
            pl.display_name AS line_display_name
        """

        if has_defect_filter:
            select_fields += """,
                MIN(od.defect_type) AS defect_type,
                MIN(od.i_ribbon) AS i_ribbon,
                MIN(od.stringa) AS stringa,
                MIN(od.ribbon_lato) AS ribbon_lato,
                MIN(od.s_ribbon) AS s_ribbon,
                MIN(od.extra_data) AS extra_data
            """

        query = f"""
        SELECT {select_fields}
        FROM productions p
        JOIN objects o ON p.object_id = o.id
        JOIN stations s ON p.station_id = s.id
        {join_sql}
        LEFT JOIN production_lines pl ON s.line_id = pl.id
        {where_sql}
        GROUP BY p.id
        {order_clause}
        LIMIT %s
        """
        params.append(limit)

        assert mysql_connection is not None
        with mysql_connection.cursor() as cursor:
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()

        return {"results": rows}

    except Exception as e:
        logging.error(f"Search API Error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.post("/api/export_objects")
def export_objects(background_tasks: BackgroundTasks, data: dict = Body(...)):
    id_moduli = data.get("id_moduli", [])
    filters = data.get("filters", [])

    global mysql_connection
    if not mysql_connection:
        return JSONResponse(status_code=500, content={"error": "MySQL connection not available"})

    try:
        export_data = {
            "id_moduli": id_moduli,
            "filters": filters,
            "moduli_rows": [],
            "objects": [],
            "productions": [],
            "stations": [],
            "production_lines": [],
            "object_defects": []
        }

        with mysql_connection.cursor() as cursor:
            # Get objects by id_modulo
            format_strings = ','.join(['%s'] * len(id_moduli))
            cursor.execute(f"SELECT * FROM objects WHERE id_modulo IN ({format_strings})", id_moduli)
            export_data["objects"] = cursor.fetchall()

            # Get all productions related to those objects
            object_ids = [o["id"] for o in export_data["objects"]]
            if not object_ids:
                return {"status": "ok", "filename": None}  # No data to export

            format_strings = ','.join(['%s'] * len(object_ids))
            cursor.execute(f"""
                SELECT * FROM productions
                WHERE object_id IN ({format_strings})
                ORDER BY start_time DESC
            """, object_ids)
            export_data["productions"] = cursor.fetchall()

            # Get stations
            cursor.execute("SELECT * FROM stations")
            export_data["stations"] = cursor.fetchall()

            # Get lines
            cursor.execute("SELECT * FROM production_lines")
            export_data["production_lines"] = cursor.fetchall()

            # Get defects
            prod_ids = [p["id"] for p in export_data["productions"]]
            if prod_ids:
                format_strings = ','.join(['%s'] * len(prod_ids))
                cursor.execute(f"""
                    SELECT od.*, d.category
                    FROM object_defects od
                    JOIN defects d ON od.defect_id = d.id
                    WHERE od.production_id IN ({format_strings})
                """, prod_ids)
                export_data["object_defects"] = cursor.fetchall()

    except Exception as e:
        logging.error(f"❌ Error during export: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})
    
    print('export Data: %s' % export_data)

    # ✅ Pass the entire export_data dict to all sheet functions
    filename = export_full_excel(export_data)
    background_tasks.add_task(clean_old_exports, max_age_hours=2)

    return {"status": "ok", "filename": filename}

# --- Download Route ---
@app.get("/api/download_export/{filename}")
def download_export(filename: str):
    filepath = os.path.join(EXPORT_DIR, filename)
    if os.path.exists(filepath):
        return FileResponse(filepath, filename=filename,
                            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        raise HTTPException(status_code=404, detail="File not found")

# ---------------- MAIN ----------------
if __name__ == "__main__":
   uvicorn.run(app, host="0.0.0.0", port=8000)